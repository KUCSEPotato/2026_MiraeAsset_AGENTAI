from __future__ import annotations

import asyncio
from datetime import UTC, date, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy.dialects import postgresql

from app.agent.service import _metric_resolutions
from app.data.cleaning import (
    canonical_mirae_sale_flag,
    canonical_subscription_status,
)
from app.data.database import DatabaseSettings
from app.data.v2_rebuild import (
    METRIC_FIELDS,
    _atomic_index,
    _date,
    _decimal,
    _etp_sale_status,
    _etp_insufficient_reasons,
    _etp_trading_status,
)
from app.data.v2_version import CANONICAL_V2_TRANSFORMER_VERSION

from app.data.metric_capabilities import (
    EVALUATION_DATA_CUTOFF,
    MetricCapabilityRegistry,
    PREF01_AUM,
    PREF01_EXPENSE,
    PREF02_AUM,
    PREF02_EXPENSE,
    PRBD_CREDIT_RATING,
    PRBD_PURCHASABLE_BOND,
    PREF01_ONE_DAY_RETURN,
    PREF01_ONE_MONTH_RETURN,
    PREF01_THREE_MONTH_RETURN,
    PREF01_SIX_MONTH_RETURN,
    PREF01_ONE_YEAR_RETURN,
    PREF01_YEAR_TO_DATE_RETURN,
    PREF02_ONE_YEAR_RETURN,
    ISHARES_SCOPED_ONE_YEAR_RETURN,
    PRFD_SHARE_CLASS_ONE_YEAR_RETURN,
    CROSS_PRODUCT_RETURN_CONTRACTS,
)
from app.domain.models import (
    CanonicalEntity,
    ExecutionContext,
    QueryPlan,
    QueryOperation,
    QueryStep,
    ResolvedQuery,
    RetrievalRecord,
    RetrievalSource,
    StepExecutionResult,
    StepExecutionStatus,
    ValidationResult,
)
from app.entity.lookup import StaticEntityLookup
from app.entity.resolver import RegistryEntityResolver
from app.evidence.answer import DeterministicEvidenceAnswerGenerator
from app.evidence.serializer import serialize_evidence_bundle
from app.execution.transforms import InternalTransformExecutor
from app.graph.config import GraphSettings
from app.ontology.loader import OntologyLoader
from app.ontology.rdf_service import RDFOntologyService
from app.ontology.runtime_mapping import TeamOntologyRuntimeMapping
from app.planning.coordinator import QueryPlanner
from app.planning.exceptions import UnsupportedQuerySemanticsError
from app.planning.metadata import RoutingMetadataRegistry
from app.planning.routing import FastRoutingChecker
from app.planning.rule_router import DeterministicRuleRouter
from app.planning.supervisor import DeterministicSupervisorPlanner
from app.planning.validator import StructuredQueryPlanValidator
from app.query.analyzer import RuleBasedQueryAnalyzer
from app.retrieval.rdb_v2 import (
    CanonicalV2FieldRegistry,
    CanonicalV2QueryCompiler,
    CanonicalV2SnapshotSelector,
    V2SnapshotSelection,
)
from app.search.config import SearchSettings
from tests.evidence_helpers import make_bundle, make_evidence


def test_canonical_v2_transformer_version_defaults_and_overrides(
    monkeypatch,
) -> None:
    test_url = "postgresql+psycopg://user@localhost/version_test"
    assert DatabaseSettings(test_url).v2_transformer_version == (
        CANONICAL_V2_TRANSFORMER_VERSION
    )
    assert SearchSettings().v2_transformer_version == CANONICAL_V2_TRANSFORMER_VERSION
    assert GraphSettings().v2_transformer_version == CANONICAL_V2_TRANSFORMER_VERSION
    assert CanonicalV2SnapshotSelector(
        snapshot_date="2026-08-24"
    )._transformer_version == CANONICAL_V2_TRANSFORMER_VERSION

    monkeypatch.setenv("DATABASE_URL", test_url)
    monkeypatch.setenv("CANONICAL_V2_TRANSFORMER_VERSION", "historical-version")
    assert DatabaseSettings.from_env().v2_transformer_version == "historical-version"
    assert SearchSettings.from_env().v2_transformer_version == "historical-version"
    assert GraphSettings.from_env().v2_transformer_version == "historical-version"
    assert CanonicalV2SnapshotSelector(
        snapshot_date="2026-08-24", transformer_version="historical-version"
    )._transformer_version == "historical-version"


def _ontology() -> RDFOntologyService:
    fields = CanonicalV2FieldRegistry().canonical_fields
    return RDFOntologyService(
        OntologyLoader(
            Path("ontology"), known_canonical_fields=fields, version="team-v1"
        ).load()
    )


def _planner() -> QueryPlanner:
    metadata = RoutingMetadataRegistry()
    return QueryPlanner(
        routing_checker=FastRoutingChecker(metadata),
        rule_router=DeterministicRuleRouter(),
        supervisor_planner=DeterministicSupervisorPlanner(),
        plan_validator=StructuredQueryPlanValidator(metadata),
    )


async def _plan(question: str):
    parsed = await RuleBasedQueryAnalyzer().analyze(question)
    grounded = await _ontology().ground(ResolvedQuery(parsed_query=parsed))
    return parsed, grounded, await _planner().create_plan(grounded)


def test_source_metric_contracts_are_source_scoped() -> None:
    assert PREF01_AUM.dataset == "PREF01N001"
    assert PREF01_AUM.currency == "KRW"
    assert PREF02_AUM.dataset == "PREF02N001"
    assert PREF02_AUM.currency == "USD"
    assert not PREF01_AUM.cross_dataset_comparability
    assert not PREF02_AUM.cross_dataset_comparability
    assert not PREF01_EXPENSE.sort_capability
    assert not PREF02_EXPENSE.sort_capability
    assert "scale" in PREF01_EXPENSE.disabled_reason
    assert PRBD_CREDIT_RATING.filter_capability
    assert PRBD_CREDIT_RATING.scale == "CREDIT_RATING_V1"
    assert PRBD_PURCHASABLE_BOND.entity_grain == "Bond"
    assert PRBD_PURCHASABLE_BOND.filter_capability
    assert PREF01_ONE_YEAR_RETURN.sort_capability
    assert PREF01_ONE_YEAR_RETURN.exact_period == "1Y"
    assert not PREF02_ONE_YEAR_RETURN.sort_capability
    assert PRFD_SHARE_CLASS_ONE_YEAR_RETURN.entity_grain == "FundShareClass"
    assert EVALUATION_DATA_CUTOFF == date(2026, 8, 24)
    assert ISHARES_SCOPED_ONE_YEAR_RETURN.sort_capability
    assert ISHARES_SCOPED_ONE_YEAR_RETURN.value_basis == (
        "issuer-published NAV total return"
    )
    assert all(not item.runtime_enabled for item in CROSS_PRODUCT_RETURN_CONTRACTS)
    assert {item.comparability for item in CROSS_PRODUCT_RETURN_CONTRACTS} == {
        "PARTIAL", "NO",
    }


def test_credit_rating_is_an_explicit_non_lexical_order() -> None:
    order = MetricCapabilityRegistry.credit_rating_order
    assert order["AAA"] > order["AA+"] > order["AA0"] > order["AA-"]
    assert order["AA-"] > order["A+"]
    assert "AAAA" not in order


@pytest.mark.parametrize(
    ("question", "field", "direction", "top_n"),
    [
        ("국내 ETF 중 순자산이 높은 상위 7개", "product.aum", "desc", 7),
        ("국내 ETF 중 순자산이 낮은 하위 4개", "product.aum", "asc", 4),
        (
            "국내 ETF 중 수익률이 높은 상위 3개 알려줘",
            "product.one_year_return",
            "desc",
            3,
        ),
        (
            "위험이 낮은 채권형 상품을 비교해줘",
            "product.risk_grade",
            "asc",
            None,
        ),
    ],
)
def test_comparative_phrases_use_one_generic_structured_pipeline(
    question: str, field: str, direction: str, top_n: int | None
) -> None:
    parsed, grounded, plan = asyncio.run(_plan(question))
    assert parsed.sort[0].direction == direction
    assert grounded.grounded_sort[0].canonical_field == field
    assert (
        parsed.result_limit.value if parsed.result_limit is not None else None
    ) == top_n
    inputs = plan.steps[0].inputs
    assert inputs["sort_operations"] == [
        {"semantic_metric_key": field, "direction": direction}
    ]
    assert inputs["comparison_contracts"][0]["sort_capability"] is True


def test_generic_return_is_resolvable_underspecification_with_disclosure() -> None:
    _, _, plan = asyncio.run(
        _plan("국내 ETF 중 수익률이 높은 상위 3개 알려줘")
    )
    resolution = plan.steps[0].inputs["comparison_contracts"][0][
        "metric_resolution"
    ]
    assert resolution == {
        "status": "RESOLVABLE_UNDERSPECIFICATION",
        "policy": "RETURN.default_period=1Y",
        "requested_phrase": "수익률",
        "resolved_metric": "product.one_year_return",
        "metric": "RETURN",
        "period": "1Y",
        "period_source": "DEFAULT_POLICY",
        "disclosure": (
            "기간이 별도로 지정되지 않아 1년 수익률을 기준으로 비교했습니다."
        ),
    }
    assert _metric_resolutions(plan) == [resolution]

    evidence = make_evidence(
        field="product.one_year_return",
        value="12.4",
        metadata={
            "real_rdb": True,
            "comparison_contracts": plan.steps[0].inputs[
                "comparison_contracts"
            ],
        },
    )
    serialized = serialize_evidence_bundle(make_bundle([evidence]))
    assert '"metric":"RETURN"' in serialized
    assert '"period":"1Y"' in serialized
    assert '"period_source":"DEFAULT_POLICY"' in serialized


def test_comparison_without_metric_remains_true_unresolved_semantics() -> None:
    question = "ETF를 비교해줘"
    parsed = asyncio.run(RuleBasedQueryAnalyzer().analyze(question))
    intent_constraint = next(
        item for item in parsed.semantic_constraints
        if item.semantic_type.value == "intent"
    )
    assert intent_constraint.payload["ambiguity_class"] == "TRUE_AMBIGUITY"
    assert intent_constraint.unsupported_reason == (
        "true_ambiguity:comparison_metric_missing"
    )
    with pytest.raises(UnsupportedQuerySemanticsError) as caught:
        asyncio.run(_plan(question))
    assert any(
        reason.startswith("unsupported_constraint:")
        for reason in caught.value.reasons
    )


def test_manager_product_context_creates_a_grounded_manager_relation() -> None:
    async def run():
        parsed = await RuleBasedQueryAnalyzer().analyze(
            "미래에셋 상품 중 순자산이 낮은 국내 ETF는?"
        )
        resolved = await RegistryEntityResolver(
            StaticEntityLookup([
                CanonicalEntity(
                    canonical_id="ORG:MIRAE-ASSET",
                    entity_type="management_company",
                    official_name="미래에셋자산운용",
                    aliases=["미래에셋"],
                )
            ])
        ).resolve(parsed)
        grounded = await _ontology().ground(resolved)
        return parsed, resolved, grounded

    parsed, resolved, grounded = asyncio.run(run())
    assert parsed.entities[0].entity_type == "management_company"
    assert resolved.resolved_entities[0].canonical_id == "ORG:MIRAE-ASSET"
    relation = grounded.grounded_relations[0]
    assert relation.raw_text == "상품 중"
    assert relation.canonical_relation == "managedBy"
    assert relation.target_value == "ORG:MIRAE-ASSET"


def test_expense_ranking_is_grounded_then_rejected_by_metric_contract() -> None:
    parsed = asyncio.run(
        RuleBasedQueryAnalyzer().analyze(
            "미래에셋 상품 중 운용보수가 낮은 국내 ETF는?"
        )
    )
    grounded = asyncio.run(_ontology().ground(ResolvedQuery(parsed_query=parsed)))
    assert grounded.grounded_sort[0].canonical_field == "product.expense_ratio"
    with pytest.raises(UnsupportedQuerySemanticsError) as caught:
        asyncio.run(_plan("국내 ETF 중 운용보수가 낮은 상품을 알려줘"))
    assert "unsupported_comparison:expense_ratio_scale_unverified" in (
        caught.value.reasons
    )


def test_explicit_threshold_is_a_filter_not_a_ranking_adjective() -> None:
    parsed = asyncio.run(
        RuleBasedQueryAnalyzer().analyze(
            "국내 ETF 중 운용보수가 0.2% 이하인 상품"
        )
    )
    assert parsed.sort == []
    assert len(parsed.filters) == 1
    assert parsed.filters[0].field == "expense_ratio"
    assert parsed.filters[0].operator.value == "lte"


def test_unsupported_explicit_return_period_is_not_replaced_by_default() -> None:
    parsed = asyncio.run(
        RuleBasedQueryAnalyzer().analyze(
            "국내 ETF 중 2년 수익률이 높은 상품을 알려줘"
        )
    )
    assert parsed.unparsed_material_spans
    with pytest.raises(UnsupportedQuerySemanticsError) as caught:
        asyncio.run(_plan(parsed.original_question))
    assert "unparsed_material_clause" in caught.value.reasons


def test_risk_order_compiles_ontology_sequence_and_stable_ties() -> None:
    _, _, plan = asyncio.run(
        _plan("위험이 낮은 채권형 상품을 비교해줘")
    )
    snapshot = V2SnapshotSelection(
        snapshot_date=date(2026, 8, 24), generation="260824",
        ontology_version="merged-optical-1.4",
        snapshot_ids=("PRBD", "PREF01", "PREF02", "PRFD"),
        dataset_ids=("PRBD", "PREF01", "PREF02", "PRFD"),
    )
    compiled = CanonicalV2QueryCompiler(
        CanonicalV2FieldRegistry(), default_limit=10
    ).compile(plan.steps[0], snapshot)
    sql = str(compiled.statement.compile(
        dialect=postgresql.dialect(), compile_kwargs={"literal_binds": True}
    ))
    assert compiled.ranking_applied
    assert "RISK_GRADE_6' THEN 1" in sql
    assert "RISK_GRADE_1' THEN 6" in sql
    assert 'entity_id COLLATE "C" ASC' in sql


def test_default_metric_disclosure_is_rendered_from_evidence() -> None:
    evidence = make_evidence(
        field="product.one_year_return",
        value="12.4",
        metadata={
            "real_rdb": True,
            "comparison_contracts": [{
                "metric_resolution": {
                    "disclosure": (
                        "기간이 별도로 지정되지 않아 1년 수익률을 "
                        "기준으로 비교했습니다."
                    )
                }
            }],
        },
    )
    answer = asyncio.run(
        DeterministicEvidenceAnswerGenerator().generate(
            "수익률이 높은 ETF",
            make_bundle([evidence]),
            ValidationResult(answerable=True),
        )
    )
    assert answer.startswith(
        "기간이 별도로 지정되지 않아 1년 수익률을 기준으로 비교했습니다."
    )


def test_percent_return_observation_is_not_rendered_as_percentage_points() -> None:
    evidence = make_evidence(
        field="product.one_year_return",
        value="12.4",
        metadata={"real_rdb": True, "metric_unit": "PERCENT"},
    )
    answer = asyncio.run(
        DeterministicEvidenceAnswerGenerator().generate(
            "1년 수익률",
            make_bundle([evidence]),
            ValidationResult(answerable=True),
        )
    )

    assert "12.4%" in answer
    assert "퍼센티지 포인트" not in answer


@pytest.mark.parametrize(
    ("question", "canonical_field", "metric_code", "period"),
    [
        ("국내 ETF 중 1D 수익률 상위 3개 알려줘", "product.one_day_return", "ONE_DAY_RETURN", "1D"),
        ("국내 ETF 중 1개월 수익률 상위 3개 알려줘", "product.one_month_return", "ONE_MONTH_RETURN", "1M"),
        ("국내 ETF 중 3개월 수익률 상위 3개 알려줘", "product.three_month_return", "THREE_MONTH_RETURN", "3M"),
        ("국내 ETF 중 6개월 수익률 상위 3개 알려줘", "product.six_month_return", "SIX_MONTH_RETURN", "6M"),
        ("국내 ETF 중 1년 수익률 상위 3개 알려줘", "product.one_year_return", "ONE_YEAR_RETURN", "1Y"),
        ("국내 ETF 중 YTD 수익률 상위 3개 알려줘", "product.year_to_date_return", "YEAR_TO_DATE_RETURN", "YTD"),
        ("국내 ETF 중 연초 이후 수익률 상위 3개 알려줘", "product.year_to_date_return", "YEAR_TO_DATE_RETURN", "YTD"),
    ],
)
def test_domestic_return_periods_use_one_structured_pipeline(
    question: str,
    canonical_field: str,
    metric_code: str,
    period: str,
) -> None:
    parsed, grounded, plan = asyncio.run(_plan(question))
    assert parsed.semantic_coverage == "complete"
    assert parsed.result_limit.value == 3
    assert grounded.grounded_sort[0].canonical_field == canonical_field
    inputs = plan.steps[0].inputs
    assert inputs["sort_operations"] == [{
        "semantic_metric_key": canonical_field,
        "direction": "desc",
    }]
    contract = inputs["comparison_contracts"][0]
    assert contract["metric"] == metric_code
    assert contract["dataset"] == "PREF01N001"
    assert contract["exact_period"] == period
    assert contract["metric_resolution"]["period"] == period
    assert contract["metric_resolution"]["period_source"] == "EXPLICIT_QUERY"


@pytest.mark.parametrize(
    ("question", "canonical_field", "metric_code", "period", "top_n"),
    [
        (
            "국내 ETF 중 최근 1년 수익률이 높은 상위 3개를 알려줘",
            "product.one_year_return",
            "ONE_YEAR_RETURN",
            "1Y",
            3,
        ),
        (
            "국내 ETF 최근 1년 수익률 상위 3개",
            "product.one_year_return",
            "ONE_YEAR_RETURN",
            "1Y",
            3,
        ),
        (
            "국내 ETF 1년 수익률 높은 3개",
            "product.one_year_return",
            "ONE_YEAR_RETURN",
            "1Y",
            3,
        ),
        (
            "국내 ETF 중 최근 6개월 수익률이 높은 상위 5개를 알려줘",
            "product.six_month_return",
            "SIX_MONTH_RETURN",
            "6M",
            5,
        ),
        (
            "국내 ETF 중 최근 6개월 수익률이 높은 상위 3개를 알려줘",
            "product.six_month_return",
            "SIX_MONTH_RETURN",
            "6M",
            3,
        ),
        (
            "국내 ETF 중 최근 3개월 수익률 상위 3개",
            "product.three_month_return",
            "THREE_MONTH_RETURN",
            "3M",
            3,
        ),
        (
            "국내 ETF 중 최근 1개월 수익률 상위 3개",
            "product.one_month_return",
            "ONE_MONTH_RETURN",
            "1M",
            3,
        ),
        (
            "국내 ETF 중 오늘 수익률 상위 3개",
            "product.one_day_return",
            "ONE_DAY_RETURN",
            "1D",
            3,
        ),
        (
            "국내 ETF 중 올해 수익률 상위 3개",
            "product.year_to_date_return",
            "YEAR_TO_DATE_RETURN",
            "YTD",
            3,
        ),
        (
            "국내 ETF 중 올해 수익률이 높은 상위 3개를 알려줘",
            "product.year_to_date_return",
            "YEAR_TO_DATE_RETURN",
            "YTD",
            3,
        ),
        (
            "국내 ETF 중 YTD 수익률 상위 3개",
            "product.year_to_date_return",
            "YEAR_TO_DATE_RETURN",
            "YTD",
            3,
        ),
    ],
)
def test_natural_return_period_phrases_are_explicit_structured_rankings(
    question: str,
    canonical_field: str,
    metric_code: str,
    period: str,
    top_n: int,
) -> None:
    parsed, grounded, plan = asyncio.run(_plan(question))

    assert parsed.semantic_coverage == "complete"
    assert parsed.unparsed_material_spans == []
    assert parsed.result_limit is not None
    assert parsed.result_limit.value == top_n
    assert parsed.sort[0].direction == "desc"
    assert grounded.grounded_sort[0].canonical_field == canonical_field
    contract = plan.steps[0].inputs["comparison_contracts"][0]
    assert contract["metric"] == metric_code
    assert contract["exact_period"] == period
    assert contract["metric_resolution"]["period"] == period
    assert contract["metric_resolution"]["period_source"] == "EXPLICIT_QUERY"


def test_pref01_return_contracts_are_independent_and_comparison_ready() -> None:
    contracts = (
        PREF01_ONE_DAY_RETURN,
        PREF01_ONE_MONTH_RETURN,
        PREF01_THREE_MONTH_RETURN,
        PREF01_SIX_MONTH_RETURN,
        PREF01_ONE_YEAR_RETURN,
        PREF01_YEAR_TO_DATE_RETURN,
    )
    assert [item.source_field for item in contracts] == [
        "du_er_1d", "du_er_1m", "du_er_3m", "du_er_6m", "du_er_1y",
        "du_er_ytd",
    ]
    assert [item.exact_period for item in contracts] == [
        "1D", "1M", "3M", "6M", "1Y", "YTD",
    ]
    assert all(item.sort_capability for item in contracts)
    assert all(item.unit == "PERCENT" for item in contracts)
    assert all(item.scale == "SOURCE_PERCENT" for item in contracts)
    assert all(not item.cross_dataset_comparability for item in contracts)

    runtime_mapping = TeamOntologyRuntimeMapping()
    rdb_fields = CanonicalV2FieldRegistry()
    for contract in contracts:
        source_metric = METRIC_FIELDS["PREF01N001"][contract.source_field]
        assert source_metric[:3] == (
            contract.metric, contract.unit, contract.scale
        )
        ontology_metric = runtime_mapping.metric(
            f"return.{contract.exact_period}"
        )
        assert ontology_metric is not None
        assert ontology_metric.canonical_field == contract.canonical_field
        assert rdb_fields.field(contract.canonical_field).semantic_key == (
            contract.metric
        )


def test_period_specific_return_compiles_allowlisted_metric_code() -> None:
    _, _, plan = asyncio.run(
        _plan("국내 ETF 중 3개월 수익률 상위 3개 알려줘")
    )
    snapshot = V2SnapshotSelection(
        snapshot_date=date(2026, 8, 24), generation="260824",
        ontology_version="merged-optical-1.4",
        snapshot_ids=("PREF01",), dataset_ids=("PREF01",),
    )
    compiled = CanonicalV2QueryCompiler(
        CanonicalV2FieldRegistry(), default_limit=10
    ).compile(plan.steps[0], snapshot)
    sql = str(compiled.statement.compile(
        dialect=postgresql.dialect(), compile_kwargs={"literal_binds": True}
    ))
    assert "THREE_MONTH_RETURN" in sql
    assert "du_er_3m" not in sql
    assert "ORDER BY" in sql


@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ("판매중", "OPEN_FOR_SUBSCRIPTION"),
        ("판매완료", "CLOSED_FOR_SUBSCRIPTION"),
        (None, None),
        ("", None),
        ("그밖의값", None),
    ],
)
def test_subscription_status_normalization(raw, expected) -> None:
    assert canonical_subscription_status(raw) == expected


@pytest.mark.parametrize(
    ("raw", "expected"),
    [("Y", True), ("N", False), (None, None), ("", None), ("?", None)],
)
def test_mirae_sale_flag_preserves_unknown(raw, expected) -> None:
    assert canonical_mirae_sale_flag(raw) is expected


@pytest.mark.parametrize(
    "question",
    [
        "현재 미래에셋에서 가입할 수 있는 공모펀드",
        "지금 추가매수 가능한 펀드",
        "미래에셋에서 판매 중인 공모펀드",
    ],
)
def test_fund_subscription_queries_compile_to_three_fact_rule(question) -> None:
    _, _, plan = asyncio.run(_plan(question))
    inputs = plan.steps[0].inputs
    assert inputs["result_grain"] == "fund_share_class"
    assert any(
        item["canonical_field"] == "product.current_fund_subscription_eligible"
        for item in inputs["filters"]
    )
    snapshot = V2SnapshotSelection(
        snapshot_date=date(2026, 8, 24), generation="260824",
        ontology_version="merged-optical-1.4",
        snapshot_ids=("PRFD01N001:test",), dataset_ids=("PRFD01N001",),
    )
    compiled = CanonicalV2QueryCompiler(
        CanonicalV2FieldRegistry(), default_limit=100
    ).compile(plan.steps[0], snapshot)
    sql = str(compiled.statement.compile(
        dialect=postgresql.dialect(), compile_kwargs={"literal_binds": True}
    ))
    assert "OFFERING_TYPE" in sql
    assert "OPEN_FOR_SUBSCRIPTION" in sql
    assert "is_sold_by_mirae_asset" in sql
    assert "liquidation" not in sql.casefold()
    assert "redemption" not in sql.casefold()


def test_fund_freshness_is_separate_from_subscription_status() -> None:
    _, _, plan = asyncio.run(
        _plan("추가매수 가능한 펀드 중 최신 기준가가 있는 상품")
    )
    fields = {item["canonical_field"] for item in plan.steps[0].inputs["filters"]}
    assert fields == {
        "product.current_fund_subscription_eligible",
        "product.latest_fund_price_available",
    }


@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ("1", "AVAILABLE_FOR_SALE"),
        ("0", "NOT_AVAILABLE_FOR_SALE"),
        (None, "UNKNOWN"),
        ("", "UNKNOWN"),
    ],
)
def test_etp_sale_status_normalization_preserves_unknown(raw, expected) -> None:
    assert _etp_sale_status(raw) == expected


@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ("0", "TRADING_ACTIVE"),
        ("1", "TRADING_HALTED"),
        (None, "UNKNOWN"),
        ("", "UNKNOWN"),
    ],
)
def test_etp_trading_status_normalization_preserves_unknown(raw, expected) -> None:
    assert _etp_trading_status(raw) == expected


def test_etp_sentinel_dates_are_not_real_lifecycle_dates() -> None:
    assert _date("00000000") is None
    assert _date("10001231") is None
    assert _date("99991231") is None
    assert _date("20260821") == date(2026, 8, 21)


def test_etp_insufficient_reasons_only_cover_core_availability_inputs() -> None:
    refinitiv_only_missing = {
        "pd_sale_yn": "1", "pd_tr_yn": "0", "pd_lstg_dt": "20260101",
        "pd_isin_cd": None, "pd_ric": None, "pd_ticker": None,
    }
    assert _etp_insufficient_reasons(refinitiv_only_missing) == ()
    assert _etp_insufficient_reasons({
        **refinitiv_only_missing, "pd_tr_yn": None,
    }) == ("TRADING_STATUS_MISSING",)
    assert _etp_insufficient_reasons({
        **refinitiv_only_missing, "pd_lstg_dt": "10001231",
    }) == ("LISTING_START_DATE_MISSING",)


def test_etp_source_metric_and_insufficient_counts_preserve_missingness() -> None:
    material = Path("material")
    # This validation intentionally runs only where both authoritative source
    # workbooks are provisioned; repository-only CI does not contain material/.
    source_workbooks = {
        prefix: sorted(material.glob(f"**/{prefix.casefold()}_data.xlsx"))
        for prefix in ("PREF01N001", "PREF02N001")
    }
    missing_sources = [
        prefix for prefix, candidates in source_workbooks.items() if not candidates
    ]
    if missing_sources:
        pytest.skip(
            "authoritative ETP source workbook(s) are not provisioned under "
            "material/**/<source>_data.xlsx: "
            + ", ".join(missing_sources)
        )

    counts = {"volume": 0, "zero_volume": 0, "missing_foreign_volume": 0}
    insufficient = {"PREF01N001": 0, "PREF02N001": 0}
    for prefix in insufficient:
        path = source_workbooks[prefix][0]
        workbook = load_workbook(path, read_only=True, data_only=True)
        rows = workbook["data"].iter_rows(values_only=True)
        header = [str(value).strip() for value in next(rows)]
        for values in rows:
            row = dict(zip(header, values, strict=False))
            volume = _decimal(row.get("ru_mkt_volume"))
            if volume is None:
                if prefix == "PREF02N001":
                    counts["missing_foreign_volume"] += 1
            else:
                counts["volume"] += 1
                counts["zero_volume"] += prefix == "PREF02N001" and volume == 0
            insufficient[prefix] += bool(_etp_insufficient_reasons(row))
        workbook.close()

    assert counts == {
        "volume": 7_799,
        "zero_volume": 91,
        "missing_foreign_volume": 14,
    }
    # Domestic raw has four core-invalid rows; one is quarantined during rebuild.
    assert insufficient == {"PREF01N001": 4, "PREF02N001": 14}


def test_pref01_return_source_contract_profile() -> None:
    candidates = sorted(
        Path("material").glob("**/pref01n001_data.xlsx")
    )
    schema_candidates = sorted(
        Path("material").glob("**/pref01n001_schema.xlsx")
    )
    if not candidates or not schema_candidates:
        pytest.skip("authoritative PREF01 source/schema workbooks are not provisioned")

    expected = {
        "du_er_1d": (1_585, 1_544),
        "du_er_1m": (1_584, 1_543),
        "du_er_3m": (1_553, 1_512),
        "du_er_6m": (1_486, 1_434),
        "du_er_1y": (1_416, 1_321),
        "du_er_ytd": (1_477, 1_422),
    }
    schema = load_workbook(schema_candidates[0], read_only=True, data_only=True)
    schema_rows = list(schema.active.iter_rows(values_only=True))
    schema_by_field = {
        str(row[1]).strip(): row for row in schema_rows[1:] if row[1] is not None
    }
    assert {
        field: (schema_by_field[field][2], schema_by_field[field][4])
        for field in expected
    } == {
        "du_er_1d": ("numeric(28,2)", "수익률_1D"),
        "du_er_1m": ("numeric(28,2)", "수익률_1M"),
        "du_er_3m": ("numeric(28,2)", "수익률_3M"),
        "du_er_6m": ("numeric(28,2)", "수익률_6M"),
        "du_er_1y": ("numeric(28,2)", "수익률_1Y"),
        "du_er_ytd": ("numeric(28,2)", "수익률_YTD"),
    }
    assert schema_by_field["du_upt_dt"][4] == "일간갱신일자"
    schema.close()

    workbook = load_workbook(candidates[0], read_only=True, data_only=True)
    rows = workbook.active.iter_rows(values_only=True)
    header = [str(value).strip() for value in next(rows)]
    counts = {field: [0, 0] for field in expected}
    for values in rows:
        row = dict(zip(header, values, strict=False))
        observed_on = _date(row.get("du_upt_dt"))
        assert observed_on is None or observed_on <= EVALUATION_DATA_CUTOFF
        for field in expected:
            if _decimal(row.get(field)) is None:
                continue
            counts[field][0] += 1
            counts[field][1] += observed_on is not None
    workbook.close()
    assert {field: tuple(value) for field, value in counts.items()} == expected


@pytest.mark.parametrize(
    ("question", "universe", "expected_region"),
    [
        ("현재 구매 가능한 국내 ETF", ["DomesticETF"], None),
        ("현재 판매 중인 국내 ETN", ["DomesticETN"], None),
        ("현재 구매 가능한 국내 ETP", ["DomesticETP"], None),
        ("최신 가격이 있는 해외 ETF", ["ForeignETF"], None),
        ("국내 주식에 투자하는 해외 ETF", ["ForeignETF"], "국내"),
        ("미국 주식에 투자하는 국내 ETF", ["DomesticETF"], "미국"),
        ("중국 시장을 추종하는 국내 ETN", ["DomesticETN"], "중국"),
    ],
)
def test_etp_universe_words_do_not_erase_explicit_exposure_regions(
    question, universe, expected_region
) -> None:
    parsed = asyncio.run(RuleBasedQueryAnalyzer().analyze(question))
    assert parsed.product_universe is not None
    assert parsed.product_universe.operands == universe
    region_filters = [item for item in parsed.filters if item.field == "region"]
    assert [item.value for item in region_filters] == (
        [] if expected_region is None else [expected_region]
    )


def test_foreign_index_sentence_sentinels_are_not_atomic_indices() -> None:
    assert not _atomic_index("Index is not provided by Management Company")
    assert not _atomic_index("Index is not available on Lipper Database")


@pytest.mark.parametrize(
    ("question", "expected_fields", "expected_universe"),
    [
        (
            "\ud604\uc7ac \uad6c\ub9e4 \uac00\ub2a5\ud55c \uad6d\ub0b4 ETF",
            {"product.current_etp_sale_eligible"},
            ["DomesticETF"],
        ),
        (
            "\ud604\uc7ac \ud310\ub9e4 \uc911\uc778 \uad6d\ub0b4 ETN",
            {"product.current_etp_sale_eligible"},
            ["DomesticETN"],
        ),
        (
            "\uac70\ub798\uc815\uc9c0\uac00 \uc544\ub2cc ETF",
            {"product.etp_trading_status"},
            None,
        ),
        (
            "\uc0c1\uc7a5 \uc885\ub8cc\ub41c ETP\ub294 \uc81c\uc678\ud574\uc918",
            {"product.etp_listing_ended"},
            ["DomesticETP", "ForeignETP"],
        ),
        (
            "\ucd5c\uc2e0 \uac00\uaca9\uc774 \uc788\ub294 \ud574\uc678 ETF",
            {"product.latest_etp_price_available"},
            ["ForeignETF"],
        ),
        (
            "\uad6c\ub9e4 \uac00\ub2a5\ud558\uc9c0\ub9cc \uac00\uaca9\uc774 \uc624\ub798\ub41c \ud574\uc678 ETF",
            {"product.current_etp_sale_eligible", "product.stale_etp_price_warning"},
            ["ForeignETF"],
        ),
        (
            "\uc815\ubcf4\uac00 \ubd80\uc871\ud574 \ucd94\ucc9c\ud558\uae30 \uc5b4\ub824\uc6b4 ETP",
            {"product.etp_insufficient_info"},
            ["DomesticETP", "ForeignETP"],
        ),
        (
            "\ud604\uc7ac \uad6c\ub9e4 \uac00\ub2a5\ud55c ETP \uc911 \ucd5c\uc2e0 \uac00\uaca9\uc774 \uc788\ub294 \uc0c1\ud488",
            {"product.current_etp_sale_eligible", "product.latest_etp_price_available"},
            ["DomesticETP", "ForeignETP"],
        ),
    ],
)
def test_etp_availability_queries_compile_to_distinct_filters(
    question, expected_fields, expected_universe
) -> None:
    _, _, plan = asyncio.run(_plan(question))
    inputs = plan.steps[0].inputs
    fields = {item["canonical_field"] for item in inputs["filters"]}
    assert expected_fields <= fields
    if expected_universe is not None:
        assert inputs["product_universe"] == {
            "operation": "UNION",
            "operands": expected_universe,
        }


def test_etp_strict_candidate_compiles_as_sale_and_latest_price_conjunction() -> None:
    _, _, plan = asyncio.run(
        _plan("\ud604\uc7ac \uad6c\ub9e4 \uac00\ub2a5\ud55c ETP \uc911 \ucd5c\uc2e0 \uac00\uaca9\uc774 \uc788\ub294 \uc0c1\ud488")
    )
    snapshot = V2SnapshotSelection(
        snapshot_date=date(2026, 8, 24), generation="260824",
        ontology_version="merged-optical-1.4",
        snapshot_ids=("PREF01N001:test", "PREF02N001:test"),
        dataset_ids=("PREF01N001", "PREF02N001"),
    )
    compiled = CanonicalV2QueryCompiler(
        CanonicalV2FieldRegistry(), default_limit=100
    ).compile(plan.steps[0], snapshot)
    sql = str(compiled.statement.compile(
        dialect=postgresql.dialect(), compile_kwargs={"literal_binds": True}
    ))
    assert "current_etp_sale_eligible" in sql
    assert "latest_etp_price_available" in sql
    assert "etp_listing_ended = false" not in sql


def test_completed_fund_exclusion_is_subscription_only() -> None:
    _, _, plan = asyncio.run(_plan("판매완료 펀드는 제외해줘"))
    item = plan.steps[0].inputs["filters"][0]
    assert item["canonical_field"] == "product.subscription_status"
    assert item["canonical_value"] == "SubscriptionStatus.CLOSED_FOR_SUBSCRIPTION"
    assert item["raw"]["operator"] == "ne"


def test_official_bond_semantics_are_structured() -> None:
    _, _, plan = asyncio.run(
        _plan("현재 판매 가능한 원화채권 중 AA- 이상 종목 알려줘")
    )
    inputs = plan.steps[0].inputs
    assert inputs["product_types"] == ["FinancialProduct.Bond"]
    fields = {item["canonical_field"]: item for item in inputs["filters"]}
    assert fields["product.currency"]["canonical_value"] == "KRW"
    assert fields["product.credit_rating"]["raw"]["operator"] == "gte"
    assert fields["product.credit_rating"]["canonical_value"] == "AA-"
    assert fields["product.current_sale_available"]["canonical_value"] is True
    assert inputs["ordered_comparisons"] == [
        {
            "semantic_field": "product.credit_rating",
            "operator": "gte",
            "value": "AA-",
        }
    ]
    assert {item["metric"] for item in inputs["comparison_contracts"]} == {
        "CREDIT_RATING_ORDER",
        "ORGANIZER_PURCHASABLE_BOND",
    }


def test_aum_top_n_uses_pref02_contract() -> None:
    _, _, plan = asyncio.run(
        _plan("미국 증시에 상장된 주식형 ETF 중 순자산이 큰 상품 3개")
    )
    inputs = plan.steps[0].inputs
    assert inputs["sort_operations"] == [
        {"semantic_metric_key": "product.aum", "direction": "desc"}
    ]
    assert inputs["top_n"] == {"value": 3}
    assert inputs["comparison_contracts"][0]["dataset"] == "PREF02N001"
    assert inputs["limit"] == 3


def test_one_year_return_top10_is_exact_period_and_source_scoped() -> None:
    parsed, _, plan = asyncio.run(
        _plan("국내 ETF 중 연 수익률 기준 TOP10 알려줘")
    )
    assert parsed.product_universe.operands == ["DomesticETF"]
    inputs = plan.steps[0].inputs
    assert inputs["sort_operations"] == [
        {"semantic_metric_key": "product.one_year_return", "direction": "desc"}
    ]
    assert inputs["top_n"] == {"value": 10}
    assert inputs["comparison_contracts"][0]["metric"] == "ONE_YEAR_RETURN"
    assert inputs["comparison_contracts"][0]["dataset"] == "PREF01N001"


@pytest.mark.parametrize(
    ("question", "operands"),
    [
        ("ETF와 공모펀드 알려줘", ["ETF", "PublicFund"]),
        (
            "국내/해외 ETF와 공모펀드 알려줘",
            ["DomesticETF", "ForeignETF", "PublicFund"],
        ),
    ],
)
def test_validated_product_universe_union_is_explicit(
    question: str, operands: list[str]
) -> None:
    parsed, _, plan = asyncio.run(_plan(question))
    assert parsed.product_universe.operands == operands
    assert plan.steps[0].inputs["product_universe"] == {
        "operation": "UNION",
        "operands": operands,
    }


def test_public_fund_return_is_not_silently_promoted_or_dropped() -> None:
    question = "국내/해외 ETF와 공모펀드 중 1년 수익률이 높은 순으로 10개 알려줘"
    with pytest.raises(UnsupportedQuerySemanticsError) as caught:
        asyncio.run(_plan(question))
    assert (
        "unsupported_comparison:public_fund_return_1Y_not_comparable_at_fund_grain"
        in caught.value.reasons
    )


@pytest.mark.parametrize("security", ["삼성전자", "SK하이닉스"])
def test_holdings_boundary_is_structured_without_production_fact_fabrication(
    security: str,
) -> None:
    parsed = asyncio.run(
        RuleBasedQueryAnalyzer().analyze(
            f"{security}를 보유한 국내/해외 ETF와 공모펀드를 "
            "연 수익률 기준 TOP10 알려줘"
        )
    )
    assert parsed.product_universe.operands == [
        "DomesticETF", "ForeignETF", "PublicFund"
    ]
    assert len(parsed.relations) == 1
    relation = parsed.relations[0]
    assert relation.raw_text == "보유한"
    assert relation.target_type == "Organization"
    assert relation.target_value == security
    assert parsed.sort[0].field == "연 수익률"
    assert parsed.result_limit.value == 10

    # C1 prepares the relation boundary but does not invent ontology mappings
    # or production holdings facts.
    with pytest.raises(UnsupportedQuerySemanticsError):
        asyncio.run(_plan(parsed.original_question))


@pytest.mark.parametrize(
    ("word", "direction"),
    [("오름차순", "asc"), ("내림차순", "desc"), ("ASC", "asc"), ("DESC", "desc")],
)
def test_explicit_sort_direction_is_preserved(word: str, direction: str) -> None:
    _, _, plan = asyncio.run(
        _plan(f"미국 증시에 상장된 ETF 중 순자산 기준 {word} 상품 3개")
    )
    assert plan.steps[0].inputs["sort_operations"] == [
        {"semantic_metric_key": "product.aum", "direction": direction}
    ]


@pytest.mark.parametrize(
    "question",
    [
        "순자산이 큰 ETF 3개",
        "미국 증시에 상장된 ETF 중 총보수가 낮은 순으로 3개",
        "상위 3개 ETF",
        "미국 증시에 상장된 ETF 중 순자산이 큰 상품 1001개",
        "현재 판매 가능한 원화채권 중 AAAA 이상 종목 알려줘",
    ],
)
def test_unsupported_comparisons_fail_before_retrieval(question: str) -> None:
    with pytest.raises(UnsupportedQuerySemanticsError):
        asyncio.run(_plan(question))


@pytest.mark.parametrize(
    ("question", "reason"),
    [
        (
            "순자산이 큰 ETF 3개",
            "unsupported_comparison:aum_scope_spans_or_cannot_exclude_incompatible_sources",
        ),
        (
            "미국 증시에 상장된 ETF 중 총보수가 낮은 순으로 3개",
            "unsupported_comparison:expense_ratio_scale_unverified",
        ),
        (
            "현재 판매 가능한 원화채권 중 AAAA 이상 종목 알려줘",
            "unsupported_comparison:invalid_credit_rating",
        ),
    ],
)
def test_unsupported_comparison_exposes_safe_reason(
    question: str, reason: str
) -> None:
    with pytest.raises(UnsupportedQuerySemanticsError) as caught:
        asyncio.run(_plan(question))
    assert reason in caught.value.reasons


def test_compiler_uses_allow_listed_sqlalchemy_ranking_and_stable_tie_break() -> None:
    _, _, plan = asyncio.run(
        _plan("미국 증시에 상장된 주식형 ETF 중 순자산이 큰 상품 3개")
    )
    snapshot = V2SnapshotSelection(
        snapshot_date=date(2026, 8, 24),
        generation="260824",
        ontology_version="merged-optical-1.3",
        snapshot_ids=("pref01", "pref02", "prbd", "prfd"),
        dataset_ids=("PREF01N001", "PREF02N001", "PRBD01N001", "PRFD01N001"),
    )
    compiled = CanonicalV2QueryCompiler(
        CanonicalV2FieldRegistry(), default_limit=100
    ).compile(plan.steps[0], snapshot)
    sql = str(
        compiled.statement.compile(
            dialect=postgresql.dialect(), compile_kwargs={"literal_binds": True}
        )
    )
    assert "metric_observations" in sql
    assert "PREF02N001" in sql
    assert "ORDER BY" in sql and "DESC" in sql and "COLLATE \"C\"" in sql
    assert "LIMIT 3" in sql
    assert "미국 증시에" not in sql
    assert compiled.ranking_applied


def test_bond_compiler_ignores_quantity_and_excludes_lifecycle_end_facts() -> None:
    _, _, plan = asyncio.run(
        _plan("현재 판매 가능한 원화채권 중 AA- 이상 종목 알려줘")
    )
    snapshot = V2SnapshotSelection(
        snapshot_date=date(2026, 8, 24),
        generation="260824",
        ontology_version="merged-optical-1.3",
        snapshot_ids=("prbd",),
        dataset_ids=("PRBD01N001",),
    )
    compiled = CanonicalV2QueryCompiler(
        CanonicalV2FieldRegistry(), default_limit=100
    ).compile(plan.steps[0], snapshot)
    sql = str(
        compiled.statement.compile(
            dialect=postgresql.dialect(), compile_kwargs={"literal_binds": True}
        )
    )
    assert "BOND_DELISTING_DATE" in sql
    assert "BOND_LISTING_END_DATE" in sql
    assert "buyable_quantity" not in sql.casefold()
    assert "CURRENT_SALE_AVAILABILITY" not in sql
    assert "CREDIT_RATING_ORDER" in sql
    assert compiled.result_grain.value == "financial_product"


@pytest.mark.parametrize(
    ("question", "expected_filters"),
    [
        (
            "현재 구매 가능한 채권",
            {"product.current_bond_purchase_eligible": True},
        ),
        (
            "현재 구매 가능한 장내채권",
            {
                "product.current_bond_purchase_eligible": True,
                "product.bond_market_presence": "EXCHANGE_TRADED",
            },
        ),
        (
            "현재 구매 가능한 장외채권",
            {
                "product.current_bond_purchase_eligible": True,
                "product.bond_market_presence": "OTC",
            },
        ),
        (
            "미래에셋 판매조건이 있는 채권",
            {"product.has_sale_lot": True},
        ),
        (
            "매매단가와 수익률이 제공된 장외채권",
            {
                "product.bond_market_presence": "OTC",
                "product.has_trade_price_and_buy_yield_sale_lot": True,
            },
        ),
        (
            "판매 LOT은 없지만 구매 가능한 채권",
            {
                "product.current_bond_purchase_eligible": True,
                "product.has_sale_lot": False,
            },
        ),
        (
            "하나의 종목에 여러 판매조건이 있는 채권",
            {"product.has_multiple_sale_lots": True},
        ),
        (
            "상장폐지 또는 리스팅 종료 채권 제외",
            {"product.current_bond_purchase_eligible": True},
        ),
    ],
)
def test_bond_query_contract_is_structured_and_compiles_at_bond_grain(
    question: str, expected_filters: dict[str, object]
) -> None:
    parsed, _, plan = asyncio.run(_plan(question))
    assert parsed.unparsed_material_spans == []
    inputs = plan.steps[0].inputs
    assert inputs["product_types"] == ["FinancialProduct.Bond"]
    assert {
        item["canonical_field"]: item["canonical_value"]
        for item in inputs["filters"]
    } == expected_filters

    snapshot = V2SnapshotSelection(
        snapshot_date=date(2026, 8, 24),
        generation="260824",
        ontology_version="merged-optical-1.4",
        snapshot_ids=("prbd",),
        dataset_ids=("PRBD01N001",),
    )
    compiled = CanonicalV2QueryCompiler(
        CanonicalV2FieldRegistry(), default_limit=100
    ).compile(plan.steps[0], snapshot)
    sql = str(compiled.statement.compile(
        dialect=postgresql.dialect(), compile_kwargs={"literal_binds": True}
    ))
    assert compiled.result_grain.value == "financial_product"
    assert "buyable_quantity" not in sql.casefold()
    if "product.current_bond_purchase_eligible" in expected_filters:
        assert "BOND_DELISTING_DATE" in sql
        assert "BOND_LISTING_END_DATE" in sql
        assert "maturity_date" not in sql
        assert "remaining_days" not in sql
    if "product.bond_market_presence" in expected_filters:
        assert "pd_exg_mkt" in sql
        assert "EXISTS" in sql
    if expected_filters.get("product.has_sale_lot") is False:
        assert "NOT (EXISTS" in sql
    if "product.has_multiple_sale_lots" in expected_filters:
        assert "count(DISTINCT" in sql
    if "product.has_trade_price_and_buy_yield_sale_lot" in expected_filters:
        assert "trade_price" in sql
        assert "buy_yield" in sql
        assert "sale_lot_price_yield_record" in sql


def test_bond_lifecycle_exclusion_is_not_product_type_or_asset_exclusion() -> None:
    parsed, _, plan = asyncio.run(
        _plan("상장폐지 또는 리스팅 종료 채권 제외")
    )
    assert parsed.product_types == ["채권"]
    assert not any(
        item.field in {"product_type", "asset_type"}
        and item.operator.value == "ne"
        for item in parsed.filters
    )
    assert plan.steps[0].inputs["product_types"] == ["FinancialProduct.Bond"]


def test_bond_market_presence_is_distinct_from_exposure_region() -> None:
    parsed, _, plan = asyncio.run(_plan("국내 장외채권 알려줘"))
    assert parsed.unparsed_material_spans == []
    fields = {
        item["canonical_field"]: item["canonical_value"]
        for item in plan.steps[0].inputs["filters"]
    }
    assert fields == {
        "product.region": "Region.KR",
        "product.bond_market_presence": "OTC",
    }


def test_allowlisted_cross_product_shape_compiles_one_global_ranking() -> None:
    step = QueryStep(
        step_id="cross-product",
        source=RetrievalSource.RDB,
        operation=QueryOperation.SEARCH_PRODUCTS,
        inputs={
            "product_types": [],
            "product_universe": {
                "operation": "UNION",
                "operands": ["DomesticETF", "ForeignETF"],
            },
            "filters": [],
            "relations": [],
            "requested_fields": [],
            "sort": [
                {
                    "raw": {"field": "1년 수익률", "direction": "desc"},
                    "canonical_field": "product.one_year_return",
                }
            ],
            "comparison_contracts": [
                {
                    **PREF01_ONE_YEAR_RETURN.as_plan_input(),
                    "dataset": "PREF01N001",
                    "datasets": ["PREF01N001", "PREF02N001"],
                    "cross_dataset_comparability": True,
                }
            ],
            "limit": 10,
            "top_n": {"value": 10},
        },
    )
    snapshot = V2SnapshotSelection(
        snapshot_date=date(2026, 8, 24), generation="260824",
        ontology_version="merged-optical-1.3",
        snapshot_ids=("pref01", "pref02"),
        dataset_ids=("PREF01N001", "PREF02N001"),
    )
    compiled = CanonicalV2QueryCompiler(
        CanonicalV2FieldRegistry(), default_limit=100
    ).compile(step, snapshot)
    sql = str(compiled.statement.compile(
        dialect=postgresql.dialect(), compile_kwargs={"literal_binds": True}
    ))
    assert sql.count("ORDER BY") >= 2  # metric scalar observation + global rank
    assert sql.count("LIMIT 10") == 1
    assert "PREF01N001" in sql and "PREF02N001" in sql


def test_ranked_intersection_preserves_all_metric_evidence_per_top_n_entity() -> None:
    rdb_step = QueryStep(
        step_id="rdb",
        source=RetrievalSource.RDB,
        operation=QueryOperation.SEARCH_PRODUCTS,
    )
    graph_step = QueryStep(
        step_id="graph",
        source=RetrievalSource.GRAPH,
        operation=QueryOperation.RELATIONSHIP_SEARCH,
    )
    rank_step = QueryStep(
        step_id="rank",
        source=RetrievalSource.INTERNAL,
        operation=QueryOperation.RANK_CANDIDATES,
        inputs={
            "limit": 2,
            "sort": [{"canonical_field": "product.one_year_return"}],
            "comparison_contracts": [{"metric": "ONE_YEAR_RETURN"}],
        },
        depends_on=["rdb", "graph"],
    )
    plan = QueryPlan(planner="supervisor", steps=[rdb_step, graph_step, rank_step])
    now = datetime.now(UTC)

    def result(step: QueryStep, records: list[RetrievalRecord], metadata=None):
        return StepExecutionResult(
            step_id=step.step_id,
            source=step.source,
            status=StepExecutionStatus.SUCCESS,
            records=records,
            retrieval_metadata=metadata or {},
            started_at=now,
            finished_at=now,
            duration_seconds=0,
        )

    rdb_records = [
        RetrievalRecord(
            source="rdb",
            source_id=f"rdb:{entity_id}:{field}",
            entity_id=entity_id,
            payload={"field": field, "value": value},
        )
        for entity_id, name, metric in (
            ("etf:1", "first", "20.0"),
            ("etf:2", "second", "10.0"),
            ("etf:3", "third", "5.0"),
        )
        for field, value in (
            ("product.name", name),
            ("product.one_year_return", metric),
        )
    ]
    graph_records = [
        RetrievalRecord(
            source="graph",
            source_id=f"graph:{entity_id}",
            entity_id=entity_id,
            payload={"field": "relation.holds", "value": "security:samsung"},
        )
        for entity_id in ("etf:1", "etf:2", "etf:3")
    ]
    context = ExecutionContext(
        plan=plan,
        step_results={
            "rdb": result(
                rdb_step,
                rdb_records,
                {
                    "rankable_total": 3,
                    "returned_count": 3,
                    "ranked_candidate_ids": ["etf:1", "etf:2", "etf:3"],
                },
            ),
            "graph": result(graph_step, graph_records),
        },
    )

    records = asyncio.run(InternalTransformExecutor().execute(rank_step, context))

    assert [record.entity_id for record in records] == [
        "etf:1", "etf:1", "etf:2", "etf:2"
    ]
    assert [record.payload["field"] for record in records] == [
        "product.name", "product.one_year_return",
        "product.name", "product.one_year_return",
    ]
    assert all(record.metadata["ranking_applied"] for record in records)
