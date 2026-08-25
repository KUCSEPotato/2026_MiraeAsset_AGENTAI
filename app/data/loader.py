from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class SourceSchema:
    columns: tuple[str, ...]
    column_types: dict[str, str]
    declared_key_columns: tuple[str, ...]


def load_source_schema(path: Path) -> SourceSchema:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = next(
            (
                name
                for name in ("Sheet1_Schema", "schema")
                if name in workbook.sheetnames
            ),
            None,
        )
        if sheet_name is None:
            raise ValueError(f"missing schema sheet: {path}")
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        first_header = tuple(str(value or "").strip() for value in (header or ()))
        if "컬럼명" not in first_header:
            legacy_header = next(rows, None)
            if legacy_header and tuple(legacy_header[:3]) == (
                "컬럼명",
                "PK/FK",
                "컬럼타입",
            ):
                header = legacy_header
        if header is None:
            raise ValueError(f"invalid schema header: {path}")

        normalized_header = tuple(str(value or "").strip() for value in header)
        if "컬럼명" not in normalized_header or not any(
            name in normalized_header for name in ("컬럼타입", "데이터타입")
        ):
            raise ValueError(f"invalid schema header: {path}")
        column_index = normalized_header.index("컬럼명")
        type_name = "컬럼타입" if "컬럼타입" in normalized_header else "데이터타입"
        type_index = normalized_header.index(type_name)
        key_index = (
            normalized_header.index("PK/FK")
            if "PK/FK" in normalized_header
            else None
        )
        columns: list[str] = []
        types: dict[str, str] = {}
        keys: list[str] = []
        for row in rows:
            column = row[column_index]
            if not column:
                continue
            name = str(column).strip()
            columns.append(name)
            types[name] = str(row[type_index] or "text").strip().casefold()
            if (
                key_index is not None
                and row[key_index]
                and "PK" in str(row[key_index]).upper()
            ):
                keys.append(name)
        return SourceSchema(tuple(columns), types, tuple(keys))
    finally:
        workbook.close()


def iter_source_rows(
    path: Path,
    schema: SourceSchema,
) -> Iterator[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = next(
            (name for name in ("datarows", "data") if name in workbook.sheetnames),
            None,
        )
        if sheet_name is None:
            raise ValueError(f"missing data sheet: {path}")
        rows = workbook[sheet_name].iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            raise ValueError(f"empty data workbook: {path}")
        headers = tuple(str(item).strip() for item in header_row if item is not None)
        if headers != schema.columns:
            missing = sorted(set(schema.columns) - set(headers))
            unexpected = sorted(set(headers) - set(schema.columns))
            raise ValueError(
                f"schema mismatch for {path.name}: "
                f"missing={missing}, unexpected={unexpected}"
            )
        for row_number, values in enumerate(rows, start=2):
            yield row_number, dict(zip(headers, values, strict=False))
    finally:
        workbook.close()
