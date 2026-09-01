from __future__ import annotations

import json
from collections import defaultdict
from dataclasses import dataclass
from enum import StrEnum
from pathlib import Path

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict

from app.external_data.holdings.providers.kodex import KodexProduct
from app.external_data.holdings.providers.kodex_catalog import KodexCatalogProduct
from app.external_data.models import EXTERNAL_KODEX_RESOLUTION_SCHEMA


class CatalogResolutionStatus(StrEnum):
    MATCHED = "MATCHED"
    AMBIGUOUS = "AMBIGUOUS"
    UNMATCHED = "UNMATCHED"


class CatalogResolution(BaseModel):
    model_config = ConfigDict(extra="forbid")

    schema_version: str = EXTERNAL_KODEX_RESOLUTION_SCHEMA
    catalog_source_id: str
    catalog_name: str
    catalog_ticker: str | None
    catalog_isin: str | None
    catalog_market: str | None
    status: CatalogResolutionStatus
    matched_by: str | None = None
    pref01_source_id: str | None = None
    pref01_name: str | None = None
    pref01_ticker: str | None = None
    pref01_isin: str | None = None

    def canonical_json(self) -> str:
        return json.dumps(
            self.model_dump(mode="json"),
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )


@dataclass(frozen=True, slots=True)
class CatalogResolutionReport:
    entries: tuple[CatalogResolution, ...]
    matched_products: tuple[KodexProduct, ...]
    matched_by_isin: int
    matched_by_ticker: int
    ambiguous: int
    unmatched: int


@dataclass(frozen=True, slots=True)
class _Pref01Row:
    source_id: str
    name: str
    ticker: str | None
    isin: str | None
    market: str | None


def resolve_catalog_against_pref01(
    products: tuple[KodexCatalogProduct, ...], pref01_data: Path,
) -> CatalogResolutionReport:
    """Resolve read-only, exact identities; no fuzzy product-name matching."""

    rows = _load_pref01(pref01_data)
    by_isin: dict[str, list[_Pref01Row]] = defaultdict(list)
    by_ticker_market: dict[tuple[str, str], list[_Pref01Row]] = defaultdict(list)
    for row in rows:
        if row.isin:
            by_isin[row.isin].append(row)
        if row.ticker and row.market:
            by_ticker_market[(row.ticker, row.market)].append(row)

    entries: list[CatalogResolution] = []
    matched: list[KodexProduct] = []
    matched_by_isin = 0
    matched_by_ticker = 0
    ambiguous = 0
    unmatched = 0
    for product in products:
        candidates: list[_Pref01Row] = []
        matched_by: str | None = None
        if product.isin:
            candidates = by_isin.get(product.isin.strip().upper(), [])
            matched_by = "ISIN" if candidates else None
        if not candidates and product.ticker and product.market:
            candidates = by_ticker_market.get(
                (product.ticker.strip().upper(), product.market.strip().upper()), []
            )
            matched_by = "TICKER_MARKET" if candidates else None
        if len(candidates) == 1:
            row = candidates[0]
            entries.append(CatalogResolution(
                catalog_source_id=product.source_id,
                catalog_name=product.name,
                catalog_ticker=product.ticker,
                catalog_isin=product.isin,
                catalog_market=product.market,
                status=CatalogResolutionStatus.MATCHED,
                matched_by=matched_by,
                pref01_source_id=row.source_id,
                pref01_name=row.name,
                pref01_ticker=row.ticker,
                pref01_isin=row.isin,
            ))
            matched.append(KodexProduct(
                source_id=product.source_id,
                name=product.name,
                ticker=row.ticker,
                isin=row.isin,
            ))
            matched_by_isin += int(matched_by == "ISIN")
            matched_by_ticker += int(matched_by == "TICKER_MARKET")
        elif len(candidates) > 1:
            ambiguous += 1
            entries.append(CatalogResolution(
                catalog_source_id=product.source_id,
                catalog_name=product.name,
                catalog_ticker=product.ticker,
                catalog_isin=product.isin,
                catalog_market=product.market,
                status=CatalogResolutionStatus.AMBIGUOUS,
                matched_by=matched_by,
            ))
        else:
            unmatched += 1
            entries.append(CatalogResolution(
                catalog_source_id=product.source_id,
                catalog_name=product.name,
                catalog_ticker=product.ticker,
                catalog_isin=product.isin,
                catalog_market=product.market,
                status=CatalogResolutionStatus.UNMATCHED,
            ))
    return CatalogResolutionReport(
        entries=tuple(sorted(entries, key=lambda item: item.catalog_source_id)),
        matched_products=tuple(sorted(matched, key=lambda item: item.source_id)),
        matched_by_isin=matched_by_isin,
        matched_by_ticker=matched_by_ticker,
        ambiguous=ambiguous,
        unmatched=unmatched,
    )


def _load_pref01(path: Path) -> list[_Pref01Row]:
    if not path.is_file():
        raise FileNotFoundError(f"authoritative PREF01 workbook not found: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = tuple(str(value).strip() if value is not None else "" for value in next(iterator))
        positions = {name: index for index, name in enumerate(headers)}
        required = {
            "pd_itm_no", "pd_abrv_nm", "pd_ticker", "pd_isin_cd",
            "pd_exg_mkt_cd", "pd_grp_no",
        }
        missing = sorted(required - positions.keys())
        if missing:
            raise ValueError(f"PREF01 identity columns missing: {missing}")
        accepted: list[_Pref01Row] = []
        for values in iterator:
            source_id = _text(values[positions["pd_itm_no"]])
            name = _text(values[positions["pd_abrv_nm"]])
            if not source_id or not name or _text(values[positions["pd_grp_no"]]) != "ETF":
                continue
            accepted.append(_Pref01Row(
                source_id=source_id,
                name=name,
                ticker=_upper(values[positions["pd_ticker"]]),
                isin=_upper(values[positions["pd_isin_cd"]]),
                market=_pref01_market(values[positions["pd_exg_mkt_cd"]]),
            ))
        return accepted
    finally:
        workbook.close()


def _text(value) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def _upper(value) -> str | None:
    normalized = _text(value)
    return normalized.upper() if normalized else None


def _pref01_market(value) -> str | None:
    normalized = _upper(value)
    if normalized == "EXG_MKT_NO_001":
        return "KRX"
    return None
