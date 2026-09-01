"""Production crawl orchestration for the reviewed TIGER source contract."""

from __future__ import annotations

import hashlib
import json
from collections import Counter
from dataclasses import dataclass
from datetime import date
from enum import StrEnum
from pathlib import Path

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict

from app.external_data.holdings.providers.tiger import (
    TigerHoldingsAdapter,
    TigerProduct,
    TigerSchemaError,
)
from app.external_data.http import TrustedHttpClient
from app.external_data.manifest import SnapshotWorkspace


TIGER_CRAWL_RESULT_SCHEMA = "external-tiger-crawl-result-v1"
_EXCLUDED_NAME_TOKENS = ("레버리지", "인버스", "커버드콜", "합성")


class TigerCrawlStatus(StrEnum):
    SUCCESS = "SUCCESS"
    FETCH_FAILED = "FETCH_FAILED"
    PARSE_FAILED = "PARSE_FAILED"
    NO_HOLDINGS = "NO_HOLDINGS"


class TigerCrawlResult(BaseModel):
    model_config = ConfigDict(extra="forbid")
    schema_version: str = TIGER_CRAWL_RESULT_SCHEMA
    product_source_id: str
    product_name: str
    product_ticker: str
    product_isin: str
    status: TigerCrawlStatus
    holding_count: int = 0
    effective_date: date | None = None
    reason: str | None = None

    def canonical_json(self) -> str:
        return json.dumps(
            self.model_dump(mode="json"), ensure_ascii=False,
            sort_keys=True, separators=(",", ":"),
        )


@dataclass(frozen=True, slots=True)
class TigerCatalogAudit:
    discovered: int
    contract_candidates: tuple[TigerProduct, ...]
    excluded: int
    unresolved_identity: int


@dataclass(frozen=True, slots=True)
class TigerProductionCrawlResult:
    catalog: TigerCatalogAudit
    first_results: tuple[TigerCrawlResult, ...]
    second_results: tuple[TigerCrawlResult, ...]
    first_holding_count: int
    second_holding_count: int
    first_semantic_checksum: str | None
    second_semantic_checksum: str | None

    @property
    def status_counts(self) -> Counter[str]:
        return Counter(item.status.value for item in self.second_results)

    @property
    def idempotent(self) -> bool:
        return (
            self.first_holding_count == self.second_holding_count
            and self.first_semantic_checksum == self.second_semantic_checksum
        )


def audit_tiger_pref01_catalog(path: Path) -> TigerCatalogAudit:
    """Resolve TIGER products from authoritative PREF01 without name matching."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(item) for item in next(iterator)]
    required = {
        "pd_abrv_nm", "pd_ticker", "pd_isin_cd", "pd_exg_mkt_nm",
        "wu_inv_rgn", "wu_inv_ast_type", "cu_strtegy",
    }
    if not required.issubset(headers):
        raise ValueError("PREF01 is missing TIGER source-contract fields")
    index = {name: headers.index(name) for name in required}
    discovered = excluded = unresolved = 0
    candidates: dict[str, TigerProduct] = {}
    for values in iterator:
        name = _text(values[index["pd_abrv_nm"]])
        if not name.upper().startswith("TIGER "):
            continue
        discovered += 1
        ticker = _ticker(values[index["pd_ticker"]])
        isin = _text(values[index["pd_isin_cd"]]).upper()
        exchange = _text(values[index["pd_exg_mkt_nm"]])
        exact_identity = len(isin) == 12 and isin.startswith("KR") and len(ticker) == 6
        safe_source_scope = (
            exchange == "유가증권"
            and _text(values[index["wu_inv_rgn"]]) == "국내"
            and _text(values[index["wu_inv_ast_type"]]) == "주식"
            and _text(values[index["cu_strtegy"]]) == "실물복제"
            and not any(token in name for token in _EXCLUDED_NAME_TOKENS)
        )
        if not exact_identity:
            unresolved += 1
            continue
        if not safe_source_scope:
            excluded += 1
            continue
        product = TigerProduct(
            source_id=isin, name=name, ticker=ticker, isin=isin,
            exchange="KRX",
        )
        product.validate()
        if isin in candidates and candidates[isin] != product:
            raise ValueError("PREF01 contains a conflicting TIGER ISIN identity")
        candidates[isin] = product
    return TigerCatalogAudit(
        discovered=discovered,
        contract_candidates=tuple(candidates[key] for key in sorted(candidates)),
        excluded=excluded,
        unresolved_identity=unresolved,
    )


async def run_tiger_production_crawl(
    client: TrustedHttpClient,
    workspace: SnapshotWorkspace,
    *,
    pref01_data: Path,
    requested_date: date,
    selected_product_isins: frozenset[str] = frozenset(),
    verify_rerun: bool = True,
) -> TigerProductionCrawlResult:
    catalog = audit_tiger_pref01_catalog(pref01_data)
    products = catalog.contract_candidates
    if selected_product_isins:
        known = {item.isin for item in products}
        if selected_product_isins - known:
            raise ValueError("selected TIGER ISIN is outside the reviewed PREF01 contract")
        products = tuple(item for item in products if item.isin in selected_product_isins)
    workspace.write_normalized_jsonl(
        category="catalog", filename="tiger_pref01_catalog.jsonl",
        schema_version="external-tiger-pref01-catalog-v1",
        canonical_rows=(
            json.dumps({
                "product_source_id": item.source_id,
                "name": item.name,
                "ticker": item.ticker,
                "isin": item.isin,
                "exchange": item.exchange,
            }, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
            for item in products
        ),
    )
    adapter = TigerHoldingsAdapter(client, workspace)
    first = await _crawl_pass(adapter, products, requested_date)
    _write_results(workspace, first, "crawl_results_pass1.jsonl")
    _write_results(workspace, first, "crawl_results.jsonl")
    first_count, first_hash = _holding_state(workspace)
    second = first
    if verify_rerun:
        successful = {
            item.product_source_id for item in first
            if item.status is TigerCrawlStatus.SUCCESS
        }
        rerun = await _crawl_pass(
            adapter, tuple(item for item in products if item.source_id in successful),
            requested_date,
        )
        by_id = {item.product_source_id: item for item in rerun}
        second = tuple(by_id.get(item.product_source_id, item) for item in first)
        _write_results(workspace, rerun, "crawl_results_pass2.jsonl")
        _write_results(workspace, second, "crawl_results.jsonl")
    second_count, second_hash = _holding_state(workspace)
    return TigerProductionCrawlResult(
        catalog=catalog,
        first_results=first,
        second_results=second,
        first_holding_count=first_count,
        second_holding_count=second_count,
        first_semantic_checksum=first_hash,
        second_semantic_checksum=second_hash,
    )


async def _crawl_pass(adapter, products, requested_date):
    results = []
    for product in products:
        try:
            result = await adapter.acquire(product, requested_date=requested_date)
            if result.source_record is None:
                status, reason = TigerCrawlStatus.FETCH_FAILED, "official source fetch failed"
            elif not result.holdings:
                status, reason = TigerCrawlStatus.NO_HOLDINGS, "official source returned no rows"
            else:
                status, reason = TigerCrawlStatus.SUCCESS, None
            count = len(result.holdings)
        except TigerSchemaError as exc:
            status, reason, count = TigerCrawlStatus.PARSE_FAILED, str(exc), 0
        results.append(TigerCrawlResult(
            product_source_id=product.source_id,
            product_name=product.name,
            product_ticker=product.ticker,
            product_isin=product.isin,
            status=status,
            holding_count=count,
            effective_date=requested_date if status is TigerCrawlStatus.SUCCESS else None,
            reason=reason,
        ))
    return tuple(sorted(results, key=lambda item: item.product_source_id))


def _write_results(workspace, rows, filename):
    workspace.write_normalized_jsonl(
        category="holdings", filename=filename,
        schema_version=TIGER_CRAWL_RESULT_SCHEMA,
        canonical_rows=[item.canonical_json() for item in rows],
    )


def _holding_state(workspace):
    path = workspace.normalized_directory("holdings") / "holdings.jsonl"
    if not path.is_file():
        return 0, None
    lines = [line for line in path.read_text(encoding="utf-8").splitlines() if line]
    return len(lines), hashlib.sha256(("\n".join(lines) + "\n").encode()).hexdigest()


def _text(value) -> str:
    return " ".join(str(value or "").strip().split())


def _ticker(value) -> str:
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return _text(value).zfill(6)
