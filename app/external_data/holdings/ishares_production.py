"""PREF02-bound production crawl orchestration for official iShares holdings."""

from __future__ import annotations

import hashlib
import json
from collections import Counter
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict

from app.external_data.holdings.providers.ishares import (
    ISharesHoldingsAdapter,
    ISharesProduct,
    ISharesSchemaError,
)
from app.external_data.http import TrustedHttpClient
from app.external_data.manifest import SnapshotWorkspace


ISHARES_HISTORICAL_DATE = date(2026, 7, 31)
ISHARES_CRAWL_RESULT_SCHEMA = "external-ishares-us-crawl-result-v1"
_PREF02_EXCHANGE_MICS = {"NAS": "XNAS", "AMX": "ARCX"}

# Source-configuration identifiers, not semantic query branches.  Each entry
# is independently reconciled to authoritative PREF02 before it can be used.
REVIEWED_ISHARES_PORTFOLIOS = {
    "IVV": "239726",
    "MCHI": "239619",
    "EWY": "239681",
    "SOXX": "239705",
    "IYW": "239522",
}


class ISharesCrawlResult(BaseModel):
    model_config = ConfigDict(extra="forbid")
    schema_version: str = ISHARES_CRAWL_RESULT_SCHEMA
    product_source_id: str
    product_name: str
    product_ticker: str
    product_isin: str
    product_exchange: str
    portfolio_id: str
    status: str
    holding_count: int = 0
    security_count: int = 0
    non_security_count: int = 0
    unsupported_count: int = 0
    effective_date: date | None = None
    reason: str | None = None

    def canonical_json(self) -> str:
        return json.dumps(
            self.model_dump(mode="json"), ensure_ascii=False,
            sort_keys=True, separators=(",", ":"),
        )


@dataclass(frozen=True, slots=True)
class PREF02ForeignETFAudit:
    foreign_etf_products: int
    with_isin: int
    with_ticker: int
    with_exchange: int
    unique_isin: int
    unique_ticker_exchange: int
    reviewed_products: tuple[ISharesProduct, ...]


@dataclass(frozen=True, slots=True)
class ISharesProductionCrawlResult:
    catalog: PREF02ForeignETFAudit
    first_results: tuple[ISharesCrawlResult, ...]
    second_results: tuple[ISharesCrawlResult, ...]
    first_holding_count: int
    second_holding_count: int
    first_semantic_checksum: str | None
    second_semantic_checksum: str | None

    @property
    def idempotent(self) -> bool:
        return (
            self.first_holding_count == self.second_holding_count
            and self.first_semantic_checksum == self.second_semantic_checksum
        )

    @property
    def status_counts(self) -> Counter[str]:
        return Counter(item.status for item in self.second_results)


def audit_pref02_foreign_etfs(path: Path) -> PREF02ForeignETFAudit:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(item).strip() for item in next(iterator)]
    required = {
        "pd_grp_no", "pd_abrv_nm", "pd_nm", "pd_isin_cd", "pd_exg_mkt_cd",
        "cu_fund_mgmt_co", "wu_inv_ast_type",
    }
    if not required.issubset(headers):
        raise ValueError("PREF02 is missing the foreign ETF identity/source fields")
    index = {name: headers.index(name) for name in required}
    rows = []
    for values in iterator:
        if _text(values[index["pd_grp_no"]]).upper() != "ETF":
            continue
        rows.append({name: _text(values[position]) for name, position in index.items()})
    by_ticker: dict[str, list[dict[str, str]]] = {}
    for row in rows:
        by_ticker.setdefault(row["pd_abrv_nm"].upper(), []).append(row)
    reviewed = []
    for ticker, portfolio_id in REVIEWED_ISHARES_PORTFOLIOS.items():
        matches = by_ticker.get(ticker, [])
        if len(matches) != 1:
            raise ValueError(f"PREF02 does not uniquely identify reviewed iShares ticker {ticker}")
        row = matches[0]
        if row["cu_fund_mgmt_co"] != "BlackRock Fund Advisors":
            raise ValueError(f"reviewed iShares product {ticker} has an unexpected provider")
        isin = row["pd_isin_cd"].upper()
        exchange = _PREF02_EXCHANGE_MICS.get(row["pd_exg_mkt_cd"].upper())
        if len(isin) != 12 or exchange is None or row["wu_inv_ast_type"] != "Equity":
            raise ValueError(f"reviewed iShares product {ticker} lacks safe PREF02 identity")
        reviewed.append(ISharesProduct(
            portfolio_id=portfolio_id,
            name=row["pd_nm"],
            ticker=ticker,
            isin=isin,
            exchange=exchange,
        ))
    isin_values = [row["pd_isin_cd"].upper() for row in rows if row["pd_isin_cd"]]
    ticker_exchange = [
        (row["pd_abrv_nm"].upper(), row["pd_exg_mkt_cd"].upper())
        for row in rows if row["pd_abrv_nm"] and row["pd_exg_mkt_cd"]
    ]
    return PREF02ForeignETFAudit(
        foreign_etf_products=len(rows),
        with_isin=sum(bool(row["pd_isin_cd"]) for row in rows),
        with_ticker=sum(bool(row["pd_abrv_nm"]) for row in rows),
        with_exchange=sum(bool(row["pd_exg_mkt_cd"]) for row in rows),
        unique_isin=len(set(isin_values)),
        unique_ticker_exchange=len(set(ticker_exchange)),
        reviewed_products=tuple(sorted(reviewed, key=lambda item: item.ticker)),
    )


async def run_ishares_production_crawl(
    client: TrustedHttpClient,
    workspace: SnapshotWorkspace,
    *,
    pref02_data: Path,
    requested_date: date = ISHARES_HISTORICAL_DATE,
    selected_tickers: frozenset[str] = frozenset(),
    verify_rerun: bool = True,
) -> ISharesProductionCrawlResult:
    audit = audit_pref02_foreign_etfs(pref02_data)
    products = audit.reviewed_products
    if selected_tickers:
        selected = {value.upper() for value in selected_tickers}
        known = {item.ticker for item in products}
        if selected - known:
            raise ValueError("selected iShares ticker is outside the reviewed PREF02 contract")
        products = tuple(item for item in products if item.ticker in selected)
    workspace.write_normalized_jsonl(
        category="catalog", filename="ishares_pref02_catalog.jsonl",
        schema_version="external-ishares-pref02-catalog-v1",
        canonical_rows=(
            json.dumps({
                "portfolio_id": item.portfolio_id,
                "product_source_id": item.source_id,
                "name": item.name,
                "ticker": item.ticker,
                "isin": item.isin,
                "exchange": item.exchange,
            }, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
            for item in products
        ),
    )
    adapter = ISharesHoldingsAdapter(client, workspace)
    first = await _crawl_pass(adapter, products, requested_date)
    _write_results(workspace, first, "crawl_results_pass1.jsonl")
    _write_results(workspace, first, "crawl_results.jsonl")
    first_count, first_hash = _holding_state(workspace)
    second = first
    if verify_rerun:
        successful = {item.product_source_id for item in first if item.status == "SUCCESS"}
        rerun = await _crawl_pass(
            adapter,
            tuple(item for item in products if item.source_id in successful),
            requested_date,
        )
        by_id = {item.product_source_id: item for item in rerun}
        second = tuple(by_id.get(item.product_source_id, item) for item in first)
        _write_results(workspace, rerun, "crawl_results_pass2.jsonl")
        _write_results(workspace, second, "crawl_results.jsonl")
    second_count, second_hash = _holding_state(workspace)
    return ISharesProductionCrawlResult(
        catalog=audit,
        first_results=first,
        second_results=second,
        first_holding_count=first_count,
        second_holding_count=second_count,
        first_semantic_checksum=first_hash,
        second_semantic_checksum=second_hash,
    )


async def _crawl_pass(adapter, products, requested_date):
    results = []
    for product in products:
        try:
            result = await adapter.acquire(product, requested_date=requested_date)
            if result.source_record is None:
                status, reason = "FETCH_FAILED", "official source fetch failed"
            elif not result.holdings:
                status, reason = "NO_HOLDINGS", "official source returned no rows"
            else:
                status, reason = "SUCCESS", None
            security = sum(
                row.position_semantic_status == "CANONICALIZABLE"
                for row in result.holdings
            )
            non_security = sum(
                row.position_semantic_status == "NON_SECURITY"
                for row in result.holdings
            )
            unsupported = len(result.holdings) - security - non_security
        except ISharesSchemaError as exc:
            status, reason = "PARSE_FAILED", str(exc)
            result = None
            security = non_security = unsupported = 0
        results.append(ISharesCrawlResult(
            product_source_id=product.source_id,
            product_name=product.name,
            product_ticker=product.ticker,
            product_isin=product.isin,
            product_exchange=product.exchange,
            portfolio_id=product.portfolio_id,
            status=status,
            holding_count=len(result.holdings) if result is not None else 0,
            security_count=security,
            non_security_count=non_security,
            unsupported_count=unsupported,
            effective_date=result.effective_date if result is not None else None,
            reason=reason,
        ))
    return tuple(sorted(results, key=lambda item: item.product_ticker))


def _write_results(workspace, rows, filename):
    workspace.write_normalized_jsonl(
        category="holdings", filename=filename,
        schema_version=ISHARES_CRAWL_RESULT_SCHEMA,
        canonical_rows=[item.canonical_json() for item in rows],
    )


def _holding_state(workspace):
    path = workspace.normalized_directory("holdings") / "holdings.jsonl"
    if not path.is_file():
        return 0, None
    lines = [line for line in path.read_text(encoding="utf-8").splitlines() if line]
    return len(lines), hashlib.sha256(("\n".join(lines) + "\n").encode()).hexdigest()


def _text(value) -> str:
    return " ".join(str(value or "").strip().split())
