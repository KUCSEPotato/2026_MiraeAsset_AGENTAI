"""Profile the local 2026-08-24 workbooks and generate the 280-column mapping."""

from __future__ import annotations

import csv
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MATERIAL = ROOT / "material" / "1.금융상품"
OUTPUT = ROOT / "ontology" / "mappings" / "column_mapping.csv"

TABLES = {
    "PRBD01N001": ("domestic_bond", "국내채권"),
    "PREF01N001": ("domestic_etp", "국내 ETF·ETN"),
    "PREF02N001": ("foreign_etp", "해외 ETF·ETN"),
    "PRFD01N001": ("fund_share_class", "펀드 클래스"),
}

IDENTIFIERS = {
    "pd_no": "ISIN", "pd_itm_no": "SOURCE_ID", "pd_isin_cd": "ISIN",
    "pd_itm_no_ma": "MA_ID", "pd_ticker": "TICKER", "pd_ric": "RIC",
    "pd_lipper_id": "LIPPER_ID", "itm_no": "SOURCE_ID", "std_itm_no": "ISIN",
    "ksd_itm_no": "KSD_ID", "rptt_ksd_itm_no": "REPRESENTATIVE_KSD_ID",
    "fss_itm_no": "FSS_ID", "mtco_itm_no": "MA_ID",
}
DATASET_IDENTIFIERS = {
    ("PREF01N001", "pd_itm_no"): "SOURCE_ID (ISIN-shaped)",
    ("PREF02N001", "pd_itm_no"): "RIC",
    ("PRFD01N001", "rptt_ksd_itm_no"): "REPRESENTATIVE_KSD_ID (FundPortfolio)",
}
DATASET_SPECIAL = {
    ("PREF02N001", "pd_itm_no_ma"): (
        "출처·관리용 메타데이터", "SourceFieldAssertion", "rawValue",
        "DataProperty", "RIC_PDF 조인키이며 독립 상품 식별자로 사용하지 않음",
    ),
}
RELATIONS = {
    "pd_pbcm": ("Issuer", "issuedBy"),
    "cu_fund_mgmt_co": ("Organization", "managedBy/issuedBy"),
    "ref_fund_mgmt_co": ("AssetManager", "managedBy"),
    "or_co_xtn_itt_cd": ("AssetManager", "managedBy"),
    "trusc_xtn_itt_cd": ("Trustee", "heldInTrustBy"),
    "cu_base_index": ("Index", "tracks"), "ref_base_index": ("Index", "tracks"),
    "bmrk_nm": ("Benchmark", "referencesBenchmark"),
    "bmrk_eng_nm": ("Benchmark", "englishName"),
    "wu_inv_rgn": ("Region", "investsInRegion"),
    "fd_ivst_rgn_desc": ("Region", "investsInRegion"),
    "wu_inv_ast_type": ("AssetType", "hasAssetType"),
    "or_attr_desc": ("AssetType", "hasAssetType"),
    "crd_grd": ("CreditRating", "hasCreditRating"),
    "pd_risk_nm": ("RiskGrade", "hasRiskGrade"),
    "zrin_fd_ivst_risk_grd_nm": ("RiskGrade", "hasRiskGrade"),
}
DIRECT = {
    "pd_nm": "productName", "itm_nm": "productName", "pd_abrv_nm": "abbreviatedName",
    "itm_abrv_nm": "abbreviatedName", "pd_eng_nm": "englishName",
    "pd_abrv_eng_nm": "englishAbbreviatedName", "itm_eng_nm": "englishName",
    "itm_eabrv_nm": "englishAbbreviatedName", "cu_strtegy": "strategyDescription",
    "cu_index_repl_mthd": "replicationMethod", "cu_lev_fector": "leverageFactor",
    "pd_grp_no": "productKind", "prvo_pbff_desc": "hasOfferingType",
}
EXCLUDED = {
    "buyable_quantity": "주최 측 규칙상 매수가능 판단에 사용할 수 없는 무효 지표; 원본 assertion만 보존",
    "pd_us_cik": "발행기관 단위일 가능성이 있는 CIK이며 상품 식별자로 사용 금지; 원본 assertion만 보존",
    "prfd_attr_search_text": "검색용 비정규화 중복 문자열; 원본 보존, 의미 관계 생성에는 사용하지 않음",
}
SPECIAL = {
    "bd_tisu_a": ("상품의 직접 속성", "Bond", "totalIssueAmount", "DataProperty", "발행 시점 총액; 통화 코드와 함께 저장"),
    "crd_grd": ("기준일이 있는 관측값", "CreditRatingObservation", "rawValue", "DataProperty", "채권 신용등급; crd_grd_dt와 결합하고 평가기관은 생성하지 않음"),
    "crd_grd_dt": ("출처·관리용 메타데이터", "CreditRatingObservation", "asOfDate", "DataProperty", "신용등급 적용일; 등급 미변경 시 과거 날짜일 수 있음"),
    "exrt_grte_ern_r": ("기준일이 있는 관측값", "YieldObservation", "metricValue", "DataProperty", "만기보장수익률; 구분 코드와 source 기준일 보존"),
    "isu_dt": ("상품의 직접 속성", "Bond", "issueDate", "DataProperty", "YYYYMMDD를 xsd:date로 검증 변환"),
    "mat_dt": ("상품의 직접 속성", "Bond", "maturityOrFirstCallDate", "DataProperty", "영구채는 1차 콜행사개시일일 수 있으므로 일반 만기로 단정하지 않음"),
    "pd_lstg_dt": ("상품의 직접 속성", "Listing", "listingStartDate", "DataProperty", "YYYYMMDD를 xsd:date로 검증 변환"),
    "pd_lste_dt": ("상품의 직접 속성", "Listing", "listingEndDate", "DataProperty", "99991231 sentinel은 날짜로 생성하지 않음"),
    "eval_price": ("기준일이 있는 관측값", "PriceObservation", "metricValue", "DataProperty", "Clean Price 성격; pd_std_info_update 기준일 연결"),
    "ndy_eval_price": ("기준일이 있는 관측값", "PriceObservation", "metricValue", "DataProperty", "익일 평가단가; 익일 관측 의미를 metricType에 보존"),
    "du_hpr": ("기준일이 있는 관측값", "PriceObservation", "metricValue", "DataProperty", "고가; 해당 일간 갱신일 연결"),
    "du_lpr": ("기준일이 있는 관측값", "PriceObservation", "metricValue", "DataProperty", "저가; 해당 일간 갱신일 연결"),
    "du_opr": ("기준일이 있는 관측값", "PriceObservation", "metricValue", "DataProperty", "시가; 해당 일간 갱신일 연결"),
    "ru_mkt_price": ("기준일이 있는 관측값", "PriceObservation", "metricValue", "DataProperty", "실시간/현재가; 원천에 시각이 없으면 날짜·시간을 추정하지 않음"),
    "du_diff_rt": ("기준일이 있는 관측값", "MetricObservation", "metricValue", "DataProperty", "가격-NAV 괴리율; 전용 기준일 또는 기준일 일치 여부와 함께 사용"),
    "du_chas_errt": ("기준일이 있는 관측값", "MetricObservation", "metricValue", "DataProperty", "추적오차율; du_chas_errt_base_dt 연결"),
    "fn_average_coupon": ("기준일이 있는 관측값", "MetricObservation", "metricValue", "DataProperty", "포트폴리오 평균 쿠폰; fn_portfolio_dt 연결"),
    "fn_average_quality": ("기준일이 있는 관측값", "MetricObservation", "rawValue", "DataProperty", "평균신용품질의 실제 값이 text/수치형 문자열이므로 의미척도 확인 전 decimal 강제 변환 금지"),
    "fn_effective_maturity": ("기준일이 있는 관측값", "MetricObservation", "metricValue", "DataProperty", "실질만기; 스키마에 단위가 없어 임의로 년을 부여하지 않음"),
    "fn_nominal_maturity": ("기준일이 있는 관측값", "MetricObservation", "metricValue", "DataProperty", "명목만기; 스키마에 단위가 없어 임의로 년을 부여하지 않음"),
    "cu_charge_rt": ("기준일이 있는 관측값", "FeeObservation", "metricValue", "DataProperty", "보수율; text 타입은 decimal 파싱 성공 시에만 관측값 생성"),
    "cu_charge_etc_rt": ("기준일이 있는 관측값", "FeeObservation", "metricValue", "DataProperty", "기타비용요율; 비교 사용은 단위 검증 후 허용"),
    "pd_dvid_pay_cnt": ("기준일이 있는 관측값", "DistributionObservation", "metricValue", "DataProperty", "연간 지급횟수; pd_dvid_base_dt 연결"),
    "exrt_rpy_r": ("기준일이 있는 관측값", "MetricObservation", "metricValue", "DataProperty", "만기상환율; 상품 고정값으로 단정하지 않고 source 기준일 보존"),
    "fd_prsv_r": ("기준일이 있는 관측값", "MetricObservation", "metricValue", "DataProperty", "보전율; 단위가 명시되지 않아 임의 환산 금지"),
    "fd_sbpr": ("기준일이 있는 관측값", "PriceObservation", "metricValue", "DataProperty", "시가평가금액; 통화 및 실제 기준일이 있을 때만 비교"),
    "bmrk_eng_nm": ("관계", "Benchmark", "benchmarkEnglishName", "DataProperty", "bmrk_nm으로 식별한 벤치마크의 영문명; 이름만으로 다른 source와 병합하지 않음"),
    "cu_lev_fector": ("상품의 직접 속성", "ExchangeTradedProduct", "leverageFactor", "DataProperty", "명시값만 decimal로 변환; 결측에서 배수를 추론하지 않음"),
    "cu_strtegy": ("상품의 직접 속성", "ExchangeTradedProduct", "strategyDescription", "DataProperty", "원문을 보존하고 요약값으로 대체하지 않음"),
    "cu_index_repl_mthd": ("상품의 직접 속성", "ExchangeTradedProduct", "replicationMethod", "DataProperty", "실제 명시값만 사용; 결측은 UNKNOWN 개체로 치환하지 않음"),
    "pd_grp_no": ("코드 또는 분류 개념", "ExchangeTradedProduct", "sourceProductKind", "DataProperty", "실제 ETF/ETN 값으로 rdf:type을 결정하고 원본 값도 보존"),
    "prvo_pbff_desc": ("코드 또는 분류 개념", "OfferingType", "hasOfferingType", "ObjectProperty", "공모/사모 명시값만 PUBLIC/PRIVATE에 매핑"),
}


@dataclass
class Profile:
    rows: int
    missing: int
    distinct: int
    examples: list[Any]
    python_types: list[str]


def profile_workbook(path: Path) -> tuple[list[str], list[Profile]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = wb["data"].iter_rows(values_only=True)
    headers = [str(value).strip() for value in next(rows)]
    missing = [0] * len(headers)
    examples: list[list[Any]] = [[] for _ in headers]
    distinct: list[set[str]] = [set() for _ in headers]
    types: list[set[str]] = [set() for _ in headers]
    count = 0
    for values in rows:
        count += 1
        for index in range(len(headers)):
            value = values[index] if index < len(values) else None
            if value is None or (isinstance(value, str) and not value.strip()):
                missing[index] += 1
                continue
            normalized = value.strip() if isinstance(value, str) else value
            distinct[index].add(str(normalized))
            types[index].add(type(normalized).__name__)
            if len(examples[index]) < 3 and normalized not in examples[index]:
                examples[index].append(normalized)
    wb.close()
    return headers, [
        Profile(count, missing[i], len(distinct[i]), examples[i], sorted(types[i]))
        for i in range(len(headers))
    ]


def read_schema(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = wb["schema"].iter_rows(values_only=True)
    header = [str(value or "").strip() for value in next(rows)]
    result = []
    for values in rows:
        if not values[header.index("컬럼명")]:
            continue
        result.append({name: str(values[i] or "").strip() for i, name in enumerate(header)})
    wb.close()
    return result


def mapping(dataset: str, column: str, description: str, data_type: str) -> tuple[str, str, str, str, str]:
    lower = column.casefold()
    if (dataset, lower) in DATASET_SPECIAL:
        return DATASET_SPECIAL[(dataset, lower)]
    if lower in EXCLUDED:
        return "현재 온톨로지 의미계층에 미반영", "SourceFieldAssertion", "rawValue", "DataProperty", EXCLUDED[lower]
    if lower in SPECIAL:
        return SPECIAL[lower]
    if lower in IDENTIFIERS:
        identifier_type = DATASET_IDENTIFIERS.get((dataset, lower), IDENTIFIERS[lower])
        return "식별자", "ProductIdentifier", "identifierValue", "DataProperty", f"identifierType={identifier_type}; 빈 값과 0/전부 0 sentinel은 identity에 사용하지 않고 원본 assertion에 보존; 형식 검증 상태를 별도 기록"
    if lower in RELATIONS:
        target, prop = RELATIONS[lower]
        role_note = "; ETF는 AssetManager/managedBy, ETN은 Issuer/issuedBy" if lower == "cu_fund_mgmt_co" else ""
        return "관계", target, prop, "ObjectProperty", f"원본 문자열/코드를 보존하고 exact 정규화만 수행; fuzzy 병합 금지{role_note}"
    if lower in DIRECT:
        return "상품의 직접 속성", "FinancialProduct", DIRECT[lower], "DataProperty", "공백 정리 외 의미 추론 금지"
    if lower in {"info_seq"}:
        return "출처·관리용 메타데이터", "BondOfferLot", "lotSequence", "DataProperty", "pd_no+pd_exg_mkt+info_base_dt+info_seq 원본 행 키의 일부"
    if lower in {"info_base_dt", "pd_std_info_update", "du_upt_dt", "cu_upt_dt", "wu_upt_dt", "ref_base_dt", "fd_daily_bas_dt"}:
        return "출처·관리용 메타데이터", "SourceRecord", "snapshotDate/asOfDate", "DataProperty", "YYYYMMDD를 xsd:date로 변환; 원본도 보존"
    is_date = lower.endswith("_dt") or lower.endswith("_base_dt") or "일자" in description or "기준일" in description
    numeric = any(token in data_type.casefold() for token in ("numeric", "double", "bigint"))
    observation_tokens = ("수익률", "금리", "가격", "단가", "기준가", "종가", "nav", "aum", "순자산", "거래", "변동성", "비율", "보수", "구성", "듀레이션", "컨벡시티", "잔존", "발행금액", "발행잔액", "주식수", "분배")
    if numeric and any(token.casefold() in description.casefold() or token.casefold() in lower for token in observation_tokens):
        return "기준일이 있는 관측값", observation_class(lower, description), "metricValue", "DataProperty", "MetricObservation으로 생성; 대응 기준일이 없으면 asOfDate를 생성하지 않음; 0과 결측을 구분"
    if is_date:
        return "출처·관리용 메타데이터", "SourceFieldAssertion", "rawValue", "DataProperty", "YYYYMMDD 형식 검증 후 날짜 의미에 맞는 개체에 연결; 불명확하면 원본만 보존"
    if lower.endswith(("_cd", "_tcd", "_gcd", "_yn", "_no", "_pcd")) or "코드" in description or "여부" in description or "구분" in description:
        return "코드 또는 분류 개념", "RawCodeValue", "codeValue", "DataProperty", "공식 코드표/설명 칼럼이 없으면 의미를 추정하지 않고 UNMAPPED_CODE로 보존"
    return "상품의 직접 속성", "SourceFieldAssertion", "rawValue", "DataProperty", "원본 값과 정규화 값을 함께 보존; 별도 의미가 확인되기 전 추론에 사용하지 않음"


def observation_class(column: str, description: str) -> str:
    text = f"{column} {description}".casefold()
    for tokens, cls in (
        (("yield", "금리", "srfc_irt"), "YieldObservation"),
        (("수익률", "yield", "ern_r"), "ReturnObservation"),
        (("변동성", "vlty"), "VolatilityObservation"),
        (("nav", "기준가"), "NAVObservation"),
        (("aum", "순자산"), "AUMObservation"),
        (("거래", "vol_", "val_"), "LiquidityObservation"),
        (("보수", "charge", "rwrd"), "FeeObservation"),
        (("분배", "dstb", "dvid"), "DistributionObservation"),
        (("구성", "cmst"), "AssetAllocationObservation"),
        (("가격", "단가", "price", "clpr", "bpr"), "PriceObservation"),
    ):
        if any(token in text for token in tokens):
            return cls
    return "MetricObservation"


def unit(description: str) -> str:
    if any(token in description for token in ("수익률", "금리", "비율", "요율", "변동성", "구성비")):
        return "PERCENT (스키마 의미 기준; 값 범위 자동 환산 금지)"
    if "일수" in description:
        return "DAY"
    if any(token in description for token in ("금액", "대금", "순자산", "AUM")):
        return "통화는 행의 통화 칼럼과 결합; 없으면 UNKNOWN"
    if any(token in description for token in ("수량", "주식수", "거래량", "횟수", "개수")):
        return "COUNT"
    return "스키마에 명시 없음"


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "원본 데이터셋", "원본 테이블", "원본 칼럼명", "스키마상의 설명", "스키마 자료형",
        "스키마 Nullable", "행 수", "결측 수", "결측률", "고유값 수", "실제 값 예시",
        "실제 Python 자료형", "분류", "대상 클래스", "대상 속성", "속성 구분", "RDF 자료형",
        "단위 또는 코드 체계", "필수 여부", "변환 규칙", "결측치 처리", "비고 및 불확실성",
    ]
    rows_out = []
    for prefix, (table, label) in TABLES.items():
        data_path = next(MATERIAL.glob(f"{prefix}_*_datarows.xlsx"))
        schema_path = next(MATERIAL.glob(f"{prefix}_*_schema.xlsx"))
        headers, profiles = profile_workbook(data_path)
        schema = read_schema(schema_path)
        if headers != [item["컬럼명"] for item in schema]:
            raise ValueError(f"schema/data header mismatch: {prefix}")
        for item, profile in zip(schema, profiles, strict=True):
            column = item["컬럼명"]
            category, target, prop, prop_kind, note = mapping(prefix, column, item["컬럼코멘트"], item["데이터타입"])
            if prop == "metricValue":
                rdf_type = "xsd:decimal"
            elif prop == "asOfDate" or prop in {"issueDate", "maturityOrFirstCallDate", "listingStartDate", "listingEndDate"}:
                rdf_type = "xsd:date"
            else:
                rdf_type = "xsd:decimal" if any(t in item["데이터타입"].casefold() for t in ("numeric", "double")) else ("xsd:integer" if "bigint" in item["데이터타입"].casefold() else "xsd:string")
            required = "조건부" if profile.missing else ("필수" if item["Nullable"].upper() == "NO" else "실제 전행 존재(스키마상 선택)")
            uncertainty = note
            if item["Nullable"].upper() == "NO" and profile.missing:
                uncertainty += f"; 스키마 Nullable=NO와 실제 결측 {profile.missing:,}건 충돌"
            rows_out.append({
                "원본 데이터셋": prefix, "원본 테이블": table, "원본 칼럼명": column,
                "스키마상의 설명": item["컬럼코멘트"], "스키마 자료형": item["데이터타입"],
                "스키마 Nullable": item["Nullable"], "행 수": profile.rows, "결측 수": profile.missing,
                "결측률": f"{profile.missing/profile.rows:.4%}", "고유값 수": profile.distinct,
                "실제 값 예시": " | ".join(str(value).replace("\n", " ")[:160] for value in profile.examples),
                "실제 Python 자료형": ",".join(profile.python_types), "분류": category,
                "대상 클래스": target, "대상 속성": prop, "속성 구분": prop_kind,
                "RDF 자료형": rdf_type, "단위 또는 코드 체계": unit(item["컬럼코멘트"]),
                "필수 여부": required, "변환 규칙": uncertainty,
                "결측치 처리": "노드를 생성하지 않으며 false/0/해당없음으로 치환하지 않음",
                "비고 및 불확실성": "실제 값 기반 프로파일; 코드 의미는 공식 코드표 없이는 미확정",
            })
    with OUTPUT.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader(); writer.writerows(rows_out)
    print(f"wrote {len(rows_out)} mappings to {OUTPUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
