"""Parse OWL, run SHACL, and verify source-backed ontology artifacts."""

from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import load_workbook
from pyshacl import validate
from rdflib import Graph, Literal, Namespace, OWL, RDF, RDFS
from rdflib.plugins.sparql import prepareQuery

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.ontology.index import FP, normalize_ontology_text  # noqa: E402
from app.ontology.loader import MANDATORY_ONTOLOGY_FILES, OntologyLoader  # noqa: E402
from app.retrieval.rdb import RDBFieldRegistry  # noqa: E402

ONTOLOGY = ROOT / "ontology"
MATERIAL = ROOT / "material" / "1.금융상품"
EX = Namespace("https://miraeasset.com/data/example/")


def shacl_conforms(data: Graph) -> tuple[bool, str]:
    shapes = Graph().parse(ONTOLOGY / "shapes.ttl", format="turtle")
    ontology = OntologyLoader(
        ONTOLOGY, known_canonical_fields=RDBFieldRegistry().canonical_fields
    ).load().graph
    conforms, _, report = validate(
        data_graph=data, shacl_graph=shapes, ont_graph=ontology,
        inference="rdfs", advanced=True, abort_on_first=False,
    )
    return bool(conforms), str(report)


def validate_alias_uniqueness(ontology: Graph) -> None:
    aliases: dict[str, set[str]] = {}
    for subject in set(ontology.subjects(FP.canonicalName, None)):
        values = {
            str(ontology.value(subject, FP.canonicalName)),
            str(subject).rsplit("#", 1)[-1],
        }
        values.update(str(value) for value in ontology.objects(subject, RDFS.label))
        values.update(str(value) for value in ontology.objects(subject, FP.alias))
        for value in values:
            aliases.setdefault(normalize_ontology_text(value), set()).add(str(subject))
    collisions = {key: values for key, values in aliases.items() if len(values) > 1}
    assert not collisions, f"ambiguous ontology aliases: {collisions}"


def assert_negative_cases(sample: Graph) -> None:
    missing_name = Graph()
    for triple in sample:
        missing_name.add(triple)
    missing_name.remove((EX["bond-KR60143NEFC6"], FP.productName, None))
    conforms, _ = shacl_conforms(missing_name)
    assert not conforms, "missing productName must fail SHACL"

    disjoint = Graph()
    for triple in sample:
        disjoint.add(triple)
    disjoint.add((EX["etf-kr-KR70000Z0003"], RDF.type, FP.ETN))
    conforms, _ = shacl_conforms(disjoint)
    assert not conforms, "ETF+ETN dual type must fail SHACL"

    collision = Graph()
    for triple in sample:
        collision.add(triple)
    collision.add((EX["duplicate-id"], RDF.type, FP.ProductIdentifier))
    for predicate, value in (
        (FP.identifierType, Literal("ISIN")),
        (FP.identifierValue, Literal("KR70000Z0003")),
        (FP.namespace, Literal("ISO-6166")),
        (FP.isPrimaryInSource, Literal(True)),
        (FP.validationStatus, Literal("FORMAT_ONLY")),
    ):
        collision.add((EX["duplicate-id"], predicate, value))
    collision.add((EX["bond-KR60143NEFC6"], FP.hasIdentifier, EX["duplicate-id"]))
    conforms, _ = shacl_conforms(collision)
    assert not conforms, "cross-product identifier collision must fail SHACL"


def validate_mapping_coverage(ontology: Graph) -> int:
    mapping_path = ONTOLOGY / "mappings" / "column_mapping.csv"
    with mapping_path.open(encoding="utf-8-sig", newline="") as handle:
        mappings = list(csv.DictReader(handle))
    keys = [(row["원본 데이터셋"], row["원본 칼럼명"]) for row in mappings]
    assert len(keys) == 280, f"expected 280 mappings, got {len(keys)}"
    assert len(keys) == len(set(keys)), "duplicate mapping rows"
    object_properties = {
        str(item).rsplit("#", 1)[-1]
        for item in ontology.subjects(RDF.type, OWL.ObjectProperty)
    }
    data_properties = {
        str(item).rsplit("#", 1)[-1]
        for item in ontology.subjects(RDF.type, OWL.DatatypeProperty)
    }
    for row in mappings:
        properties = row["대상 속성"].replace("snapshotDate/asOfDate", "snapshotDate/asOfDate").split("/")
        for prop in properties:
            expected = object_properties if row["속성 구분"] == "ObjectProperty" else data_properties
            assert prop in expected, (
                f"undeclared or wrong-kind property: {row['원본 데이터셋']}."
                f"{row['원본 칼럼명']} -> {prop}"
            )
    if not MATERIAL.is_dir():
        return len(keys)
    expected: list[tuple[str, str]] = []
    for prefix in ("PRBD01N001", "PREF01N001", "PREF02N001", "PRFD01N001"):
        schema_path = next(MATERIAL.glob(f"{prefix}_*_schema.xlsx"))
        workbook = load_workbook(schema_path, read_only=True, data_only=True)
        rows = workbook["schema"].iter_rows(values_only=True)
        header = [str(value or "").strip() for value in next(rows)]
        column_index = header.index("컬럼명")
        expected.extend(
            (prefix, str(row[column_index]).strip())
            for row in rows if row[column_index]
        )
        workbook.close()
    assert keys == expected, "mapping does not exactly cover latest schema columns"
    return len(keys)


def validate_source_samples() -> None:
    if not MATERIAL.is_dir():
        return
    expectations = {
        "PRBD01N001": ("pd_no", "KR60143NEFC6", "pd_nm", "퍼스트파이브지제팔십사차유동화전문1-14", None),
        "PREF01N001": ("pd_itm_no", "KR70000Z0003", "pd_nm", "KB RISE 바이오TOP10액티브증권상장지수투자신탁(주식)", "ETF"),
        "PREF02N001": ("pd_itm_no", "AAUA.K", "pd_nm", "Alpha Architect US Equity 3 ETF", "ETF"),
        "PRFD01N001": ("itm_no", "KR5010101611", "itm_nm", "크리스탈신종MMFE1-1비주식", "공모"),
    }
    for prefix, (key, expected, name_key, expected_name, classification) in expectations.items():
        path = next(MATERIAL.glob(f"{prefix}_*_datarows.xlsx"))
        workbook = load_workbook(path, read_only=True, data_only=True)
        rows = workbook["data"].iter_rows(values_only=True)
        header = [str(value).strip() for value in next(rows)]
        found = False
        for values in rows:
            row = dict(zip(header, values, strict=False))
            if row.get(key) == expected:
                assert row.get(name_key) == expected_name
                if classification == "ETF":
                    assert row.get("pd_grp_no") == "ETF"
                elif classification == "공모":
                    assert row.get("prvo_pbff_desc") == "공모"
                found = True
                break
        workbook.close()
        assert found, f"sample source row not found: {prefix}/{expected}"
    # ETN is a separate representative from the same domestic ETP file.
    path = next(MATERIAL.glob("PREF01N001_*_datarows.xlsx"))
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = workbook["data"].iter_rows(values_only=True)
    header = [str(value).strip() for value in next(rows)]
    assert any(
        dict(zip(header, values, strict=False)).get("pd_itm_no") == "KRG520000438"
        and dict(zip(header, values, strict=False)).get("pd_grp_no") == "ETN"
        for values in rows
    )
    workbook.close()


def main() -> None:
    loaded = OntologyLoader(
        ONTOLOGY, known_canonical_fields=RDBFieldRegistry().canonical_fields
    ).load()
    assert len(loaded.files) == len(MANDATORY_ONTOLOGY_FILES)
    validate_alias_uniqueness(loaded.graph)
    Graph().parse(ONTOLOGY / "shapes.ttl", format="turtle")
    sample = Graph().parse(ONTOLOGY / "examples" / "sample_instances.ttl", format="turtle")
    conforms, report = shacl_conforms(sample)
    assert conforms, report
    required_classes = (FP.Bond, FP.ETF, FP.ETN, FP.FundShareClass)
    assert all(any(sample.triples((None, RDF.type, cls))) for cls in required_classes)
    assert_negative_cases(sample)
    query_count = 0
    for query_path in sorted((ONTOLOGY / "queries").glob("*.rq")):
        prepareQuery(query_path.read_text(encoding="utf-8"))
        query_count += 1
    assert query_count == 5
    mapping_count = validate_mapping_coverage(loaded.graph)
    validate_source_samples()
    print(
        f"ontology_ok files={len(loaded.files)} triples={len(loaded.graph)} "
        f"sample_triples={len(sample)} mappings={mapping_count} "
        f"sparql_queries={query_count} shacl=conforms negative_cases=3/3"
    )


if __name__ == "__main__":
    main()
